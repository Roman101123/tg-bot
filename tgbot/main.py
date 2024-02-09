from aiogram import Bot, Dispatcher, executor, types
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext

from config import *
from keyboards import *
from states_groups import *
from db import Database
from constants import *
from paginators import Paginator, PaginatorDiary

from datetime import date, datetime, timedelta
import sys
from apscheduler.schedulers.asyncio import AsyncIOScheduler
import matplotlib.pyplot as plt
import math
import os
from jinja2 import Template
import json
import openpyxl
import logging
from logging.handlers import RotatingFileHandler

# зависимости
#     pip install aiogram
#     pip install apscheduler
#     pip install matplotlib
#     pip install Jinja2
#     pip install openpyxl

storage = MemoryStorage()
bot = Bot(token=token, parse_mode="HTML")
dp = Dispatcher(bot, storage=storage)
db = Database(path_to_database)
scheduler = AsyncIOScheduler(timezone="Europe/Moscow")
with open(path_to_html_template, "r", encoding="utf8") as file:
    template = Template(file.read())

def update_username(callmes: types.Message | types.CallbackQuery):
    username = db.get_username_by_chat_id(callmes.from_user.id)
    logging.debug(f"update_username | {callmes.from_user.id = }; {username = }")

    if username:
        if username[1] != callmes.from_user.username:
            logging.info(f"update_username | Изменение username. ID = {callmes.from_user.id}; Старый username = {username[1]}; Новый username = {callmes.from_user.username}")
            if callmes.from_user.username is not None:
                if db.is_user_exists_by_username(callmes.from_user.username):
                    db.update_username_in_users(callmes.from_user.username)
                    db.update_username_in_block_set_null(callmes.from_user.username)
            db.update_names_in_users(username[0], callmes.from_user.username, callmes.from_user.full_name)
            db.update_username_in_block(username[0], callmes.from_user.username)
    else:
        logging.info(f"update_username | Добавление пользователя. {callmes.from_user.id = }; {callmes.from_user.username = }")
        if callmes.from_user.username is not None:
            if db.is_user_exists_by_username(callmes.from_user.username):  
                db.update_username_in_users(callmes.from_user.username)
                db.update_username_in_block_set_null(callmes.from_user.username)
                db.add_user(callmes.from_user.id, callmes.from_user.full_name, callmes.from_user.username, datetime.now(), date.today())
            else:
                db.add_user(callmes.from_user.id, callmes.from_user.full_name, callmes.from_user.username, datetime.now(), date.today())
                db.update_user_id_in_block(callmes.from_user.username, db.get_user_id_and_datetime_with_chat_id(callmes.from_user.id)[0])
        else:
            db.add_user(callmes.from_user.id, callmes.from_user.full_name, callmes.from_user.username, datetime.now(), date.today())

async def on_startup(dp: Dispatcher):
    logging.info("🚀 Бот успешно запущен!")
    
    for admin in admins:
        await bot.send_message(admin,
                               text="🚀 <i>Бот успешно запущен!</i>\n\nВы авторизовались как <b>админ</b>.\n\nДля <b>начала работы</b> с ботом введите /start\nЧтобы получить <b>список команд</b> введи /help\nЧтобы вызвать <b>админскую панель</b> введите /admin_panel")

async def is_admin_filter(callmes: types.Message | types.CallbackQuery):
    update_username(callmes)

    ret = callmes.from_user.id in admins
    logging.debug(f"is_admin_filter | Админ - {ret}; {callmes.from_user.id = }; {callmes.from_user.username = }")

    return ret

async def is_not_block_filter(callmes: types.Message | types.CallbackQuery):
    update_username(callmes)

    if ((str(callmes.from_user.id), ) not in db.get_block_chat_ids()) and ((callmes.from_user.username, ) not in db.get_block_usernames() or callmes.from_user.username is None):
        logging.debug(f"is_not_block_filter | Не заблокирован; {callmes.from_user.id = }; {callmes.from_user.username = }")
        return True
    logging.debug(f"is_not_block_filter | Заблокирован; {callmes.from_user.id = }; {callmes.from_user.username = }")
    return False

@dp.message_handler(is_not_block_filter, commands=['start'])
async def start(message: types.Message, state: FSMContext):
    logging.info(f"{sys._getframe().f_code.co_name} | {message.chat.id = }; {message.from_user.username = }; {await state.get_data() = }")

    await state.reset_data()

    await message.answer(
        f"👋 Привет, {message.from_user.first_name}!\n\nЭтот бот поможет найти <b>калорийность продуктов</b> и <b>крутые рецепты</b>, а также вычислить <b>дневную норму калорий</b> для тебя.\nТакже ты сможешь учитывать <b>количество съеденных калорий</b>",
        reply_markup=menu_markup)
    db.update_end_message(0, message.chat.id)


@dp.message_handler(is_not_block_filter, commands=['help'])
async def help(message: types.Message, state: FSMContext):
    logging.info(f"{sys._getframe().f_code.co_name} | {message.chat.id = }; {message.from_user.username = }; {await state.get_data() = }")

    await state.reset_data()

    text = f"👋 Привет, {message.from_user.first_name}!\n\n📋 Список команд:\n<b>/start</b> - начало работы с ботом\n<b>/help</b> - список команд"
    if message.from_user.id in admins:
        text += "\n<b>/admin_panel</b> - админская панель"
    await message.answer(text)
    db.update_end_message(0, message.chat.id)


@dp.callback_query_handler(is_not_block_filter, text="product_calories")
async def product_calories(callback: types.CallbackQuery):
    logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }")

    await callback.message.edit_text(text="Что хотите <b>найти</b>?", reply_markup=product_calories_markup)


@dp.callback_query_handler(is_not_block_filter, text="products")
async def products_menu(callback: types.CallbackQuery, state: FSMContext):
    logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {await state.get_data() = }")
    
    await state.reset_data()

    await callback.message.edit_text(text="Нажмите <b>на кнопку</b>:", reply_markup=products_markup)


@dp.callback_query_handler(is_not_block_filter, lambda call: call.data.startswith("productscategories1"))
async def products_categories1(callback: types.CallbackQuery, state: FSMContext):
    logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {await state.get_data() = }")
    async with state.proxy() as data:
        if 'categories_1' in data:
            func_data = data['categories_1']
        else:
            func_data = {}

        if "paginator" in func_data:
            paginator = func_data['paginator']

            if "categories_2" in data:
                item_pages = [data['categories_2'].get(str(data_[0]), {"page": 0})['page'] for data_ in
                              paginator.get_data()]
            else:
                item_pages = [0] * len(paginator.get_data())
        else:
            paginator = Paginator(data=db.get_all_categories1(), callback_prefix="productscategories1_0",
                                  back_callback="products", type="list", width=2,
                                  height=3, item_prefix="productscategories2", second_type="products")
            func_data['paginator'] = paginator
            item_pages = [0] * len(paginator.get_data())
        func_data['page'] = int(callback.data.split("_")[-1])
        data['categories_1'] = func_data

    await callback.message.edit_text(text="Выберите <b>категорию</b>:",
                                     reply_markup=paginator.get_page_keyboard(callback.data, item_pages))


@dp.callback_query_handler(is_not_block_filter, lambda call: call.data.startswith("productscategories2"))
async def products_categories2(callback: types.CallbackQuery, state: FSMContext):
    logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {await state.get_data() = }")
    async with state.proxy() as data:
        func_data = {}
        if "categories_2" in data:
            cat_data = data[f'categories_2']
            if callback.data.split("_")[1] in cat_data:
                func_data = cat_data[callback.data.split("_")[1]]
        else:
            data['categories_2'] = {}

        if "paginator" in func_data:
            paginator = func_data["paginator"]

            if "products" in data:
                item_pages = [data['products'].get(str(data_[0]), {"page": 0})['page'] for data_ in
                              paginator.get_data()]
            else:
                item_pages = [0] * len(paginator.get_data())
        else:
            page_categories1 = data["categories_1"]["page"]
            paginator = Paginator(data=db.get_categories2(int(callback.data.split("_")[1])),
                                  callback_prefix=f"productscategories2_{callback.data.split('_')[1]}",
                                  back_callback=f"productscategories1_0_{page_categories1}", type="list", width=2,
                                  height=3,
                                  item_prefix="products", second_type="products")
            func_data["paginator"] = paginator

            item_pages = [0] * len(paginator.get_data())
        func_data['page'] = int(callback.data.split("_")[-1])
        data['categories_2'][callback.data.split("_")[1]] = func_data

    await callback.message.edit_text(text="Выберите <b>подкатегорию</b>:",
                                     reply_markup=paginator.get_page_keyboard(callback.data, item_pages))


@dp.callback_query_handler(is_not_block_filter, lambda call: call.data.startswith("products"))
async def products(callback: types.CallbackQuery, state: FSMContext):
    logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {await state.get_data() = }")
    async with state.proxy() as data:
        func_data = {}
        if "products" in data:
            products_data = data[f'products']
            if callback.data.split("_")[1] in products_data:
                func_data = products_data[callback.data.split("_")[1]]
        else:
            data['products'] = {}

        if "paginator" in func_data:
            paginator = func_data["paginator"]
        else:
            page_categories2 = data["categories_2"][str(db.get_categories_1_id(int(callback.data.split('_')[1])))][
                "page"]
            paginator = Paginator(data=db.get_products(int(callback.data.split("_")[1])),
                                  callback_prefix=f"products_{callback.data.split('_')[1]}",
                                  back_callback=f"productscategories2_{db.get_categories_1_id(int(callback.data.split('_')[1]))}_{page_categories2}",
                                  type="itemlist", width=2, height=4, item_prefix="itemproducts",
                                  category_id=callback.data.split("_")[1], second_type="products")
            func_data["paginator"] = paginator
        func_data['page'] = int(callback.data.split("_")[-1])
        data['products'][callback.data.split("_")[1]] = func_data

    await callback.message.edit_text(text="Выберите <b>продукт</b>:",
                                     reply_markup=paginator.get_page_keyboard(callback.data, [i for i in range(
                                         len(paginator.get_data()))]))


@dp.callback_query_handler(is_not_block_filter, lambda call: call.data.startswith("itemproducts"))
async def item_products(callback: types.CallbackQuery, state: FSMContext):
    db.register_activity(callback.message.chat.id, "item_products", datetime.now(), date.today())
    logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {await state.get_data() = }")
    async with state.proxy() as data:
        func_data = {}
        if "item_products" in data:
            products_data = data[f'item_products']
            if callback.data.split('_')[1] in products_data:
                func_data = products_data[callback.data.split('_')[1]]
        else:
            data['item_products'] = {}

        page_products = data["products"][callback.data.split('_')[1]]["page"]
        if "paginator" in func_data:
            paginator = func_data["paginator"]

            n_page_products = data["products"][callback.data.split('_')[1]]["paginator"].get_page_by_data_ind(
                int(callback.data.split('_')[-1]))
            if page_products != n_page_products:
                data["products"][callback.data.split('_')[1]]["page"] = n_page_products
                paginator.back_callback = f"products_{callback.data.split('_')[1]}_{n_page_products}"
            elif int(paginator.back_callback.split("_")[-1]) != n_page_products:
                paginator.back_callback = f"products_{callback.data.split('_')[1]}_{n_page_products}"
        else:
            paginator = Paginator(data=db.get_products(int(callback.data.split('_')[1])),
                                  callback_prefix=f"itemproducts_{callback.data.split('_')[1]}",
                                  back_callback=f"products_{callback.data.split('_')[1]}_{page_products}",
                                  type="items", width=1, second_type="products")
            func_data["paginator"] = paginator

        data['item_products'][callback.data.split('_')[1]] = func_data

    product = paginator.get_data()[int(callback.data.split("_")[-1])]
    await callback.message.edit_text(
        text=f"<b>{product[1]}</b><a href=\"{product[2]}\"> </a>\n\n🔥 <i>Энергетическая ценность</i> на 100 гр - <b>{product[-1]} ккал</b>\n\n<i>Пищевая ценность</i> на 100 гр:\n💪 <i>Белки</i> - <b>{product[4]} гр</b>\n🧈 <i>Жиры</i> - <b>{product[5]} гр</b>\n🏃 <i>Углеводы </i> - <b>{product[6]} гр</b>",
        reply_markup=paginator.get_page_keyboard(callback.data, option=db.is_product_favorite(callback.from_user.id, product[0])))


@dp.callback_query_handler(is_not_block_filter, text="search_products")
async def search_products(callback: types.CallbackQuery):
    logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }")
    await SearchProducts.start(callback)


class SearchProducts:
    @staticmethod
    async def cancel(message: types.Message, state: FSMContext):
        logging.debug(f"SearchProducts.cancel | {message.chat.id = }; {message.from_user.username = }; {await state.get_state() = }; {await state.get_data() = }")
        if message.text == "✖️ Отмена":
            logging.info(f"SearchProducts.cancel | Пользователь нажал \"Отмена\"; {message.chat.id = }; {message.from_user.username = }; {await state.get_state() = }; {await state.get_data() = }")
            await state.finish()

            message = await message.answer(text="🕐 Загрузка ...", reply_markup=remove_markup)
            await message.delete()

            await message.answer(text="Нажмите <b>на кнопку</b>:", reply_markup=products_markup)
            await send_mailing(message.chat.id)

            return False
        return True

    @staticmethod
    async def start(callback: types.CallbackQuery):
        logging.info(f"SearchProducts.start | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }")
        await SearchProductsStatesGroup.first()

        await callback.message.delete()
        await callback.message.answer(text="Введите <b>название продукта</b>:", reply_markup=cancel_markup)
        db.update_end_message(0, callback.message.chat.id)

    @staticmethod
    @dp.message_handler(is_not_block_filter, content_types=['text'], state=SearchProductsStatesGroup.q)
    async def get_q(message: types.Message, state: FSMContext):
        logging.info(f"SearchProducts.{sys._getframe().f_code.co_name} | {message.chat.id = }; {message.from_user.username = }; {await state.get_state() = }; {await state.get_data() = }")
        if await SearchProducts.cancel(message, state):
            await state.finish()
            await state.update_data(user_products_search={"q": message.text})

            await answer_products_search(message, state)
            await send_mailing(message.chat.id)


@dp.callback_query_handler(is_not_block_filter, lambda call: call.data.startswith("answerproductssearch"))
async def answer_products_search(callmes: types.CallbackQuery | types.Message, state: FSMContext):
    if isinstance(callmes, types.CallbackQuery):
        logging.info(f"{sys._getframe().f_code.co_name} | {callmes.message.chat.id = }; {callmes.from_user.username = }; {callmes.data = }; {await state.get_data() = }")
    else:
        logging.info(f"{sys._getframe().f_code.co_name} | {callmes.chat.id = }; {callmes.from_user.username = }; {await state.get_data() = }")
        
    async with state.proxy() as data:
        if isinstance(callmes, types.Message):
            func_data = {}
            paginator = Paginator(data=db.get_products_by_search(data["user_products_search"]['q']),
                                  callback_prefix=f"answerproductssearch_0",
                                  back_callback="products",
                                  type="itemlist", width=2, height=4, item_prefix="itemanswerproductssearch",
                                  category_id=0, second_type="products")
            func_data['paginator'] = paginator
            func_data['page'] = 0
        else:
            func_data = data['product_search']
            paginator = func_data['paginator']
            func_data['page'] = int(callmes.data.split("_")[-1])

        data['product_search'] = func_data

    if isinstance(callmes, types.Message):
        message = await callmes.answer(text="🕐 Загрузка ...", reply_markup=remove_markup)
        await message.delete()

        if paginator.get_data():
            await callmes.answer(text="Выберите <b>продукт</b>:",
                                 reply_markup=paginator.get_page_keyboard(0, [i for i in range(
                                     len(paginator.get_data()))]))
        else:
            await callmes.answer(text="❌ <i>Не удалось найти продукт</i>", reply_markup=to_products_or_menu_markup)
    else:
        await callmes.message.edit_text(text="Выберите <b>продукт</b>:",
                                        reply_markup=paginator.get_page_keyboard(callmes.data, [i for i in range(
                                            len(paginator.get_data()))]))


@dp.callback_query_handler(is_not_block_filter, lambda call: call.data.startswith("itemanswerproductssearch"))
async def item_answer_products_search(callback: types.CallbackQuery, state: FSMContext):
    db.register_activity(callback.message.chat.id, "item_answer_products_search", datetime.now(), date.today())
    logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {await state.get_data() = }")
    async with state.proxy() as data:
        if "item_product_search" in data:
            func_data = data[f'item_product_search']
        else:
            func_data = {}

        page_products_search = data["product_search"]["page"]
        if "paginator" in func_data:
            paginator = func_data['paginator']

            n_page_products_search = data["product_search"]["paginator"].get_page_by_data_ind(
                int(callback.data.split('_')[-1]))
            if page_products_search != n_page_products_search:
                data["product_search"]["page"] = n_page_products_search
                paginator.back_callback = f"answerproductssearch_0_{n_page_products_search}"
            elif int(paginator.back_callback.split("_")[-1]) != n_page_products_search:
                paginator.back_callback = f"answerproductssearch_0_{n_page_products_search}"
        else:
            paginator = Paginator(data=data['product_search']['paginator'].get_data(),
                                  callback_prefix=f"itemanswerproductssearch_0",
                                  back_callback=f"answerproductssearch_0_{page_products_search}",
                                  type="items", width=1, second_type="products")
            func_data['paginator'] = paginator

        data['item_product_search'] = func_data

    product = paginator.get_data()[int(callback.data.split("_")[-1])]
    await callback.message.edit_text(
        text=f"<b>{product[1]}</b><a href=\"{product[2]}\"> </a>\n\n🔥 <i>Энергетическая ценность</i> на 100 гр - <b>{product[-1]} ккал</b>\n\n<i>Пищевая ценность</i> на 100 гр:\n💪 <i>Белки</i> - <b>{product[4]} гр</b>\n🧈 <i>Жиры</i> - <b>{product[5]} гр</b>\n🏃 <i>Углеводы </i> - <b>{product[6]} гр</b>",
        reply_markup=paginator.get_page_keyboard(callback.data, option=db.is_product_favorite(callback.from_user.id, product[0])))


@dp.callback_query_handler(is_not_block_filter, text="recipes")
async def recipes_menu(callback: types.CallbackQuery, state: FSMContext):
    logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {await state.get_data() = }")
    
    await state.reset_data()

    await callback.message.edit_text(text="Нажмите <b>на кнопку</b>:", reply_markup=recipes_markup)


@dp.callback_query_handler(is_not_block_filter, lambda call: call.data.startswith("recipescategories"))
async def recipes_categories1(callback: types.CallbackQuery, state: FSMContext):
    logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {await state.get_data() = }")
    async with state.proxy() as data:
        if 'recipes_categories' in data:
            func_data = data['recipes_categories']
        else:
            func_data = {}

        if "paginator" in func_data:
            paginator = func_data['paginator']

            if "recipes" in data:
                item_pages = [data['recipes'].get(str(data_[0]), {"page": 0})['page'] for data_ in
                              paginator.get_data()]
            else:
                item_pages = [0] * len(paginator.get_data())
        else:
            paginator = Paginator(data=db.get_all_recipes_categories(), callback_prefix="recipescategories_0",
                                  back_callback="recipes", type="list", width=2,
                                  height=3, item_prefix="recipes", second_type="recipes")
            func_data['paginator'] = paginator
            item_pages = [0] * len(paginator.get_data())

        func_data['page'] = int(callback.data.split("_")[-1])
        data['recipes_categories'] = func_data

    await callback.message.edit_text(text="Выберите <b>категорию</b>:",
                                     reply_markup=paginator.get_page_keyboard(callback.data, item_pages))


@dp.callback_query_handler(is_not_block_filter, lambda call: call.data.startswith("recipes"))
async def recipes(callback: types.CallbackQuery, state: FSMContext):
    logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {await state.get_data() = }")
    async with state.proxy() as data:
        func_data = {}
        if "recipes" in data:
            recipes_data = data[f'recipes']
            if callback.data.split("_")[1] in recipes_data:
                func_data = recipes_data[callback.data.split("_")[1]]
        else:
            data['recipes'] = {}

        if "paginator" in func_data:
            paginator = func_data["paginator"]
        else:
            page_categories = data["recipes_categories"]["page"]
            paginator = Paginator(data=db.get_recipes(int(callback.data.split("_")[1])),
                                  callback_prefix=f"recipes_{callback.data.split('_')[1]}",
                                  back_callback=f"recipescategories_0_{page_categories}",
                                  type="itemlist", width=2, height=4, item_prefix="itemrecipes",
                                  category_id=callback.data.split("_")[1], second_type="recipes")
            func_data["paginator"] = paginator
        func_data['page'] = int(callback.data.split("_")[-1])
        data['recipes'][callback.data.split("_")[1]] = func_data

    await callback.message.edit_text(text="Выберите <b>рецепт</b>:",
                                     reply_markup=paginator.get_page_keyboard(callback.data, [i for i in range(
                                         len(paginator.get_data()))]))


@dp.callback_query_handler(is_not_block_filter, lambda call: call.data.startswith("itemrecipes"))
async def item_recipes(callback: types.CallbackQuery, state: FSMContext):
    db.register_activity(callback.message.chat.id, "item_recipes", datetime.now(), date.today())
    logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {await state.get_data() = }")
    async with state.proxy() as data:
        func_data = {}
        if "item_recipes" in data:
            recipes_data = data[f'item_recipes']
            if callback.data.split('_')[1] in recipes_data:
                func_data = recipes_data[callback.data.split('_')[1]]
        else:
            data['item_recipes'] = {}

        page_recipes = data["recipes"][callback.data.split('_')[1]]["page"]
        if "paginator" in func_data:
            paginator = func_data["paginator"]

            n_page_recipes = data["recipes"][callback.data.split('_')[1]]["paginator"].get_page_by_data_ind(
                int(callback.data.split('_')[-1]))
            if page_recipes != n_page_recipes:
                data["recipes"][callback.data.split('_')[1]]["page"] = n_page_recipes
                paginator.back_callback = f"recipes_{callback.data.split('_')[1]}_{n_page_recipes}"
            elif int(paginator.back_callback.split("_")[-1]) != n_page_recipes:
                paginator.back_callback = f"recipes_{callback.data.split('_')[1]}_{n_page_recipes}"
        else:
            paginator = Paginator(data=db.get_recipes(int(callback.data.split('_')[1])),
                                  callback_prefix=f"itemrecipes_{callback.data.split('_')[1]}",
                                  back_callback=f"recipes_{callback.data.split('_')[1]}_{page_recipes}",
                                  type="items", width=1, second_type="recipes")
            func_data["paginator"] = paginator

        data['item_recipes'][callback.data.split('_')[1]] = func_data

    recipe = paginator.get_data()[int(callback.data.split("_")[-1])]
    await callback.message.edit_text(
        text=f"<b>{recipe[1]}</b><a href=\"{recipe[2]}\"> </a>\n\n🔥 <i>Энергетическая ценность</i> на 100 гр - <b>{recipe[7]} ккал</b>\n\n<i>Пищевая ценность</i> на 100 гр:\n💪 <i>Белки</i> - <b>{recipe[4]} гр</b>\n🧈 <i>Жиры</i> - <b>{recipe[5]} гр</b>\n🏃 <i>Углеводы </i> - <b>{recipe[6]} гр</b>\n\n<b><i><a href=\"{recipe[-1]}\">Инструкция по приготовлению</a></i></b>",
        reply_markup=paginator.get_page_keyboard(callback.data, option=db.is_recipe_favorite(callback.from_user.id, recipe[0])))


@dp.callback_query_handler(is_not_block_filter, text="search_recipes")
async def search_recipes(callback: types.CallbackQuery):
    logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }")
    await SearchRecipes.start(callback)


class SearchRecipes:
    @staticmethod
    async def cancel(message: types.Message, state: FSMContext):
        logging.debug(f"SearchRecipes.cancel | {message.chat.id = }; {message.from_user.username = }; {await state.get_state() = }; {await state.get_data() = }")
        if message.text == "✖️ Отмена":
            logging.info(f"SearchRecipes.cancel | Пользователь нажал \"Отмена\"; {message.chat.id = }; {message.from_user.username = }; {await state.get_state() = }; {await state.get_data() = }")
            await state.finish()

            message = await message.answer(text="🕐 Загрузка ...", reply_markup=remove_markup)
            await message.delete()

            await message.answer(text="Нажмите <b>на кнопку</b>:", reply_markup=recipes_markup)
            await send_mailing(message.chat.id)

            return False
        return True

    @staticmethod
    async def start(callback: types.CallbackQuery):
        logging.info(f"SearchRecipes.start | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }")
        await SearchRecipesStatesGroup.first()

        await callback.message.delete()
        await callback.message.answer(text="Введите <b>название рецепта</b>:", reply_markup=cancel_markup)
        db.update_end_message(0, callback.message.chat.id)

    @staticmethod
    @dp.message_handler(is_not_block_filter, content_types=['text'], state=SearchRecipesStatesGroup.q)
    async def get_q(message: types.Message, state: FSMContext):
        logging.info(f"SearchRecipes.{sys._getframe().f_code.co_name} | {message.chat.id = }; {message.from_user.username = }; {await state.get_state() = }; {await state.get_data() = }")
        if await SearchRecipes.cancel(message, state):
            await state.finish()
            await state.update_data(user_recipes_search={"q": message.text})

            await answer_recipes_search(message, state)
            await send_mailing(message.chat.id)


@dp.callback_query_handler(is_not_block_filter, lambda call: call.data.startswith("answerrecipessearch"))
async def answer_recipes_search(callmes: types.CallbackQuery | types.Message, state: FSMContext):
    if isinstance(callmes, types.CallbackQuery):
        logging.info(f"{sys._getframe().f_code.co_name} | {callmes.message.chat.id = }; {callmes.from_user.username = }; {callmes.data = }; {await state.get_data() = }")
    else:
        logging.info(f"{sys._getframe().f_code.co_name} | {callmes.chat.id = }; {callmes.from_user.username = }; {await state.get_data() = }")

    async with state.proxy() as data:
        if isinstance(callmes, types.Message):
            func_data = {}
            paginator = Paginator(data=db.get_recipes_by_search(data["user_recipes_search"]['q']),
                                  callback_prefix=f"answerrecipessearch_0",
                                  back_callback="recipes",
                                  type="itemlist", width=2, height=4, item_prefix="itemanswerrecipessearch",
                                  category_id=0, second_type="recipes")
            func_data['paginator'] = paginator
            func_data['page'] = 0
        else:
            func_data = data['recipe_search']
            paginator = func_data['paginator']
            func_data['page'] = int(callmes.data.split("_")[-1])

        data['recipe_search'] = func_data

    if isinstance(callmes, types.Message):
        message = await callmes.answer(text="🕐 Загрузка ...", reply_markup=remove_markup)
        await message.delete()

        if paginator.get_data():
            await callmes.answer(text="Выберите <b>рецепт</b>:",
                                 reply_markup=paginator.get_page_keyboard(0, [i for i in range(
                                     len(paginator.get_data()))]))
        else:
            await callmes.answer(text="❌ <i>Не удалось найти рецепт</i>", reply_markup=to_recipes_or_menu_markup)
    else:
        await callmes.message.edit_text(text="Выберите <b>рецепт</b>:",
                                        reply_markup=paginator.get_page_keyboard(callmes.data, [i for i in range(
                                            len(paginator.get_data()))]))


@dp.callback_query_handler(is_not_block_filter, lambda call: call.data.startswith("itemanswerrecipessearch"))
async def item_answer_recipes_search(callback: types.CallbackQuery, state: FSMContext):
    db.register_activity(callback.message.chat.id, "item_answer_recipes_search", datetime.now(), date.today())
    logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {await state.get_data() = }")
    async with state.proxy() as data:
        if "item_recipes_search" in data:
            func_data = data[f'item_recipes_search']
        else:
            func_data = {}

        page_recipes_search = data["recipe_search"]["page"]
        if "paginator" in func_data:
            paginator = func_data['paginator']

            n_page_recipes_search = data["recipe_search"]["paginator"].get_page_by_data_ind(
                int(callback.data.split('_')[-1]))
            if page_recipes_search != n_page_recipes_search:
                data["recipe_search"]["page"] = n_page_recipes_search
                paginator.back_callback = f"answerrecipessearch_0_{n_page_recipes_search}"
            elif int(paginator.back_callback.split("_")[-1]) != n_page_recipes_search:
                paginator.back_callback = f"answerrecipessearch_0_{n_page_recipes_search}"
        else:
            paginator = Paginator(data=data['recipe_search']['paginator'].get_data(),
                                  callback_prefix=f"itemanswerrecipessearch_0",
                                  back_callback=f"answerrecipessearch_0_{page_recipes_search}",
                                  type="items", width=1, second_type="recipes")
            func_data['paginator'] = paginator

        data['item_recipes_search'] = func_data

    recipe = paginator.get_data()[int(callback.data.split("_")[-1])]
    await callback.message.edit_text(
        text=f"<b>{recipe[1]}</b><a href=\"{recipe[2]}\"> </a>\n\n🔥 <i>Энергетическая ценность</i> на 100 гр - <b>{recipe[7]} ккал</b>\n\n<i>Пищевая ценность</i> на 100 гр:\n💪 <i>Белки</i> - <b>{recipe[4]} гр</b>\n🧈 <i>Жиры</i> - <b>{recipe[5]} гр</b>\n🏃 <i>Углеводы </i> - <b>{recipe[6]} гр</b>\n\n<b><i><a href=\"{recipe[-1]}\">Инструкция по приготовлению</a></i></b>",
        reply_markup=paginator.get_page_keyboard(callback.data, option=db.is_recipe_favorite(callback.from_user.id, recipe[0])))


@dp.callback_query_handler(is_not_block_filter, lambda call: call.data.startswith("addfavorite"))
async def add_favorite(callback: types.CallbackQuery):
    logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }")

    type_ = callback.data.split("_")[1]
    id_ = callback.data.split("_")[-2]
    dt = datetime.now()

    if type_ == "products":
        db.add_favorite_product(callback.from_user.id, id_, dt)
        await callback.answer("➕ Продукт добавлен в избранное")
    else:
        db.add_favorite_recipe(callback.from_user.id, id_, dt)
        await callback.answer("➕ Рецепт добавлен в избранное")

    keyboard = callback.message.reply_markup.inline_keyboard
    keyboard[keyboard.index([InlineKeyboardButton(text="➕ Добавить в избранное", callback_data=callback.data)])] = [InlineKeyboardButton(text="❌ Удалить из избранного", callback_data=callback.data.replace("add", "del"))]
    
    await callback.message.edit_reply_markup(reply_markup=types.InlineKeyboardMarkup(inline_keyboard=keyboard))


@dp.callback_query_handler(is_not_block_filter, lambda call: call.data.startswith("delfavorite"))
async def del_favorite(callback: types.CallbackQuery):
    logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }")

    type_ = callback.data.split("_")[1]
    id_ = callback.data.split("_")[-2]

    if type_ == "products":
        db.del_favorite_product(callback.from_user.id, id_)
        await callback.answer("❌ Продукт удален из избранного")
    else:
        db.del_favorite_recipe(callback.from_user.id, id_)
        await callback.answer("❌ Рецепт удален из избранного")

    keyboard = callback.message.reply_markup.inline_keyboard

    keyboard[keyboard.index([InlineKeyboardButton(text="❌ Удалить из избранного", callback_data=callback.data)])] = [InlineKeyboardButton(text="➕ Добавить в избранное", callback_data=callback.data.replace("del", "add"))]
    
    await callback.message.edit_reply_markup(reply_markup=types.InlineKeyboardMarkup(inline_keyboard=keyboard))


@dp.callback_query_handler(is_not_block_filter, text="favourites")
async def favourites(callback: types.CallbackQuery, state: FSMContext):
    logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {await state.get_data() = }")

    await state.reset_data()

    await callback.message.edit_text("Выберите <b>пункт</b>:", reply_markup=favourites_markup)

@dp.callback_query_handler(is_not_block_filter, lambda call: call.data.startswith("favproducts"))
async def fav_products(callback: types.CallbackQuery, state: FSMContext):
    async with state.proxy() as data:
        logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {data.as_dict() = }")

        if "fav_products" in data:
            func_data = data["fav_products"]
        else:
            func_data = {}

        if "paginator" in func_data:
            paginator = func_data["paginator"]
        else:
            favorite_products = db.get_favorites_products(callback.from_user.id)

            if not favorite_products:
                await callback.answer("❗️ Вы ещё ничего не добавляли в избранное", show_alert=True)
                return

            paginator = Paginator(data=favorite_products,
                                  callback_prefix=f"favproducts_0",
                                  back_callback=f"favourites",
                                  type="itemlist", width=2, height=4, item_prefix="itemfavproducts",
                                  category_id=0, second_type="products")
            func_data["paginator"] = paginator

        func_data['page'] = int(callback.data.split("_")[-1])
        data['fav_products'] = func_data

    await callback.message.edit_text(text="Выберите <b>продукт</b>",
                                     reply_markup=paginator.get_page_keyboard(callback.data, [i for i in range(len(paginator.get_data()))]))

@dp.callback_query_handler(is_not_block_filter, lambda call: call.data.startswith("itemfavproducts"))
async def item_fav_products(callback: types.CallbackQuery, state: FSMContext):
    async with state.proxy() as data:
        logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {data.as_dict() = }")
        if "item_fav_products" in data:
            func_data = data[f'item_fav_products']
        else:
            func_data = {}

        page_fav_products = data["fav_products"]["page"]
        if "paginator" in func_data:
            paginator = func_data['paginator']

            n_page_fav_products = data["fav_products"]["paginator"].get_page_by_data_ind(int(callback.data.split('_')[-1]))
            if page_fav_products != n_page_fav_products:
                data["fav_products"]["page"] = n_page_fav_products
                paginator.back_callback = f"favproducts_0_{n_page_fav_products}"
            elif int(paginator.back_callback.split("_")[-1]) != n_page_fav_products:
                paginator.back_callback = f"favproducts_0_{n_page_fav_products}"
        else:
            paginator = Paginator(data=data['fav_products']['paginator'].get_data(),
                                  callback_prefix=f"itemfavproducts_0",
                                  back_callback=f"favproducts_0_{page_fav_products}",
                                  type="items", width=1, second_type="products")
            func_data['paginator'] = paginator

        data["item_fav_products"] = func_data

    product = paginator.get_data()[int(callback.data.split("_")[-1])]
    await callback.message.edit_text(
        text=f"<b>{product[1]}</b><a href=\"{product[2]}\"> </a>\n\n🔥 <i>Энергетическая ценность</i> на 100 гр - <b>{product[-1]} ккал</b>\n\n<i>Пищевая ценность</i> на 100 гр:\n💪 <i>Белки</i> - <b>{product[4]} гр</b>\n🧈 <i>Жиры</i> - <b>{product[5]} гр</b>\n🏃 <i>Углеводы </i> - <b>{product[6]} гр</b>",
        reply_markup=paginator.get_page_keyboard(callback.data, option=db.is_product_favorite(callback.from_user.id, product[0])))


@dp.callback_query_handler(is_not_block_filter, lambda call: call.data.startswith("favrecipes"))
async def fav_recipes(callback: types.CallbackQuery, state: FSMContext):
    async with state.proxy() as data:
        logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {data.as_dict() = }")

        if "fav_recipes" in data:
            func_data = data["fav_recipes"]
        else:
            func_data = {}

        if "paginator" in func_data:
            paginator = func_data["paginator"]
        else:
            favorite_recipes = db.get_favorites_recipes(callback.from_user.id)

            if not favorite_recipes:
                await callback.answer("❗️ Вы ещё ничего не добавляли в избранное", show_alert=True)
                return

            paginator = Paginator(data=favorite_recipes,
                                  callback_prefix=f"favrecipes_0",
                                  back_callback=f"favourites",
                                  type="itemlist", width=2, height=4, item_prefix="itemfavrecipes",
                                  category_id=0, second_type="recipes")
            func_data["paginator"] = paginator

        func_data['page'] = int(callback.data.split("_")[-1])
        data['fav_recipes'] = func_data

    await callback.message.edit_text(text="Выберите <b>рецепт</b>",
                                     reply_markup=paginator.get_page_keyboard(callback.data, [i for i in range(len(paginator.get_data()))]))

@dp.callback_query_handler(is_not_block_filter, lambda call: call.data.startswith("itemfavrecipes"))
async def item_fav_recipes(callback: types.CallbackQuery, state: FSMContext):
    async with state.proxy() as data:
        logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {data.as_dict() = }")
        if "item_fav_recipes" in data:
            func_data = data[f'item_fav_recipes']
        else:
            func_data = {}

        page_fav_recipes = data["fav_recipes"]["page"]
        if "paginator" in func_data:
            paginator = func_data['paginator']

            n_page_fav_recipes = data["fav_recipes"]["paginator"].get_page_by_data_ind(int(callback.data.split('_')[-1]))
            if page_fav_recipes != n_page_fav_recipes:
                data["fav_recipes"]["page"] = n_page_fav_recipes
                paginator.back_callback = f"favrecipes_0_{n_page_fav_recipes}"
            elif int(paginator.back_callback.split("_")[-1]) != n_page_fav_recipes:
                paginator.back_callback = f"favrecipes_0_{n_page_fav_recipes}"
        else:
            paginator = Paginator(data=data['fav_recipes']['paginator'].get_data(),
                                  callback_prefix=f"itemfavrecipes_0",
                                  back_callback=f"favrecipes_0_{page_fav_recipes}",
                                  type="items", width=1, second_type="recipes")
            func_data['paginator'] = paginator

        data["item_fav_recipes"] = func_data

    recipe = paginator.get_data()[int(callback.data.split("_")[-1])]
    await callback.message.edit_text(
        text=f"<b>{recipe[1]}</b><a href=\"{recipe[2]}\"> </a>\n\n🔥 <i>Энергетическая ценность</i> на 100 гр - <b>{recipe[7]} ккал</b>\n\n<i>Пищевая ценность</i> на 100 гр:\n💪 <i>Белки</i> - <b>{recipe[4]} гр</b>\n🧈 <i>Жиры</i> - <b>{recipe[5]} гр</b>\n🏃 <i>Углеводы </i> - <b>{recipe[6]} гр</b>\n\n<b><i><a href=\"{recipe[-1]}\">Инструкция по приготовлению</a></i></b>",
        reply_markup=paginator.get_page_keyboard(callback.data, option=db.is_recipe_favorite(callback.from_user.id, recipe[0])))

@dp.callback_query_handler(is_not_block_filter, lambda call: call.data.startswith("adddiary"))
async def add_to_diary(callback: types.CallbackQuery, state: FSMContext):
    logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {await state.get_data() = }")

    product_id = None
    recipe_id = None

    type_ = callback.data.split("_")[1]
    id_ = callback.data.split("_")[-2]

    if type_ == "products":
        product_id = id_
    else:
        recipe_id = id_

    await AddToDiary.start(callback, state, product_id, recipe_id, callback.message.parse_entities(), callback.message.reply_markup.inline_keyboard)

class AddToDiary:
    @staticmethod
    async def cancel(message: types.Message, state: FSMContext):
        logging.debug(f"AddToDiary.cancel | {message.chat.id = }; {message.from_user.username = }; {await state.get_state() = }; {await state.get_data() = }")
        if message.text == "✖️ Отмена":
            logging.info(f"AddToDiary.cancel | Пользователь нажал \"Отмена\"; {message.chat.id = }; {message.from_user.username = }; {await state.get_state() = }; {await state.get_data() = }")
            data = await state.get_data()
            await state.reset_state(with_data=False)

            message = await message.answer(text="🕐 Загрузка ...", reply_markup=remove_markup)
            await message.delete()
            await message.answer(data["message_text"], reply_markup=types.InlineKeyboardMarkup(inline_keyboard=data["message_keyboard"]))

            await send_mailing(message.chat.id)

            return False
        return True

    @staticmethod
    async def start(callback: types.CallbackQuery, state: FSMContext, product_id, recipe_id, message_text, message_keyboard):
        logging.info(f"AddToDiary.start | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {await state.get_state() = }; {await state.get_data() = }")
        await AddToDiaryStatesGroup.first()
        await state.update_data(product_id=product_id, recipe_id=recipe_id, message_text=message_text, message_keyboard=message_keyboard)

        if db.get_end_message(callback.from_user.id) == 0:
            await callback.message.edit_text("🍳 Выберите <b>прием пищи</b>:", reply_markup=choose_the_group_markup)
        else:
            await callback.message.delete()
            await callback.message.answer("🍳 Выберите <b>прием пищи</b>:", reply_markup=choose_the_group_markup)
            db.update_end_message(0, callback.from_user.id)

    @staticmethod
    @dp.callback_query_handler(is_not_block_filter, state=AddToDiaryStatesGroup.get_group)
    async def get_group(callback: types.CallbackQuery, state: FSMContext):
        logging.info(f"AddToDiary.{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {await state.get_state() = }; {await state.get_data() = }")
        if callback.data == "cancel":
            data = await state.get_data()
            await state.reset_state(with_data=False)

            await callback.message.edit_text(data["message_text"], reply_markup=types.InlineKeyboardMarkup(inline_keyboard=data["message_keyboard"]))

            await send_mailing(callback.message.chat.id)
        else:
            await state.update_data(group=callback.data)
            await AddToDiaryStatesGroup.next()

            await callback.message.delete()
            await callback.message.answer("⚖️ Отправьте <b>вес</b> (гр) съеденной порции в виде целого числа:", reply_markup=cancel_markup)

    @staticmethod
    @dp.message_handler(is_not_block_filter, content_types=['text'], state=AddToDiaryStatesGroup.get_volume)
    async def get_volume(message: types.Message, state: FSMContext):
        logging.info(f"AddToDiary.{sys._getframe().f_code.co_name} | {message.chat.id = }; {message.from_user.username = }; {await state.get_state() = }; {await state.get_data() = }")
        if await AddToDiary.cancel(message, state):
            try:
                volume = int(message.text)
                if not 1 <= volume <= 100000:
                    raise Exception

                await state.reset_state(with_data=False)

                data = await state.get_data()
                db.add_to_diary(data["product_id"], data["recipe_id"], message.from_user.id, volume, data["group"], date.today())

                message = await message.answer(text="🕐 Загрузка ...", reply_markup=remove_markup)
                await message.delete()
                await message.answer("✍️ <i>Прием пищи записан</i>\n\n" + data["message_text"], reply_markup=types.InlineKeyboardMarkup(inline_keyboard=data["message_keyboard"]))

                await send_mailing(message.chat.id)
            except Exception as e:
                logging.warning(f"get_volume | Error - {e}; {e.args = }")
                await message.answer(
                    "❌ <i>Неверный формат</i>.\n\n⚖️ Отправьте <b>вес</b> (гр) съеденной порции в виде целого числа ещё раз:",
                    reply_markup=cancel_markup)


@dp.callback_query_handler(is_not_block_filter, lambda call: call.data.startswith("diary"))
async def diary(callback: types.CallbackQuery):
    logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }")

    if "_" in callback.data:
        dt = callback.data.split("_")[-1]
    else:
        dt = date.today()

    items = db.get_diary(callback.from_user.id, dt)

    text = ""

    breakfast = "\n".join([f"<i>{item[2]}</i>: Калории - {round(item[6] / 100 * item[0])} кКал, Белки - {round(item[3] / 100 * item[0], 1)} гр, Жиры - {round(item[4] / 100 * item[0], 1)} гр, Углеводы - {round(item[5] / 100 * item[0], 1)} гр; Вес - {item[0]} гр" for item in items if item[1] == "breakfast"])
    if breakfast:
        text += "🍳 <b>Завтрак</b>:\n" + breakfast

    lunch = "\n".join([f"<i>{item[2]}</i>: Калории - {round(item[6] / 100 * item[0])} кКал, Белки - {round(item[3] / 100 * item[0], 1)} гр, Жиры - {round(item[4] / 100 * item[0], 1)} гр, Углеводы - {round(item[5] / 100 * item[0], 1)} гр; Вес - {item[0]} гр" for item in items if item[1] == "lunch"])
    if lunch:
        text += "\n\n🍜 <b>Обед</b>:\n" + lunch

    dinner = "\n".join([f"<i>{item[2]}</i>: Калории - {round(item[6] / 100 * item[0])} кКал, Белки - {round(item[3] / 100 * item[0], 1)} гр, Жиры - {round(item[4] / 100 * item[0], 1)} гр, Углеводы - {round(item[5] / 100 * item[0], 1)} гр; Вес - {item[0]} гр" for item in items if item[1] == "dinner"])
    if dinner:
        text += "\n\n🍝 <b>Ужин</b>:\n" + dinner

    snack = "\n".join([f"<i>{item[2]}</i>: Калории - {round(item[6] / 100 * item[0])} кКал, Белки - {round(item[3] / 100 * item[0], 1)} гр, Жиры - {round(item[4] / 100 * item[0], 1)} гр, Углеводы - {round(item[5] / 100 * item[0], 1)} гр; Вес - {item[0]} гр" for item in items if item[1] == "snack"])
    if snack:
        text += "\n\n🍎 <b>Перекус</b>:\n" + snack

    if text:
        zip_items = list(zip(*[[item[0]] + [item[3] / 100 * item[0]] + [item[4] / 100 * item[0]] + [item[5] / 100 * item[0]] + [item[6] / 100 * item[0]] for item in items]))

        sum_proteins = sum(zip_items[1])
        sum_fats = sum(zip_items[2])
        sum_carbohydrates = sum(zip_items[3])
        sum_calories = sum(zip_items[4])
        sum_volume = sum(zip_items[0])

        text += f"\n\n<b>Всего</b>: Калории - {round(sum_calories)} кКал, Белки - {round(sum_proteins, 1)} гр, Жиры - {round(sum_fats, 1)} гр, Углеводы - {round(sum_carbohydrates, 1)} гр; Вес - {sum_volume} гр"
    else:
        text = "❌ <i>Вы ещё не записывали ни один продукт или рецепт</i>"

    await callback.message.edit_text(text, reply_markup=PaginatorDiary.get_page_keyboard(dt))

@dp.callback_query_handler(is_not_block_filter, text="person_calories")
async def person_calories(callmes: types.CallbackQuery | types.Message):
    if isinstance(callmes, types.CallbackQuery):
        logging.info(f"{sys._getframe().f_code.co_name} | {callmes.message.chat.id = }; {callmes.from_user.username = }; {callmes.data = }")
        await callmes.message.edit_text(text="Что вы хотите вычислить?", reply_markup=person_calories_markup)
    else:
        logging.info(f"{sys._getframe().f_code.co_name} | {callmes.chat.id = }; {callmes.from_user.username = }")
        await callmes.answer(text="Что вы хотите вычислить?", reply_markup=person_calories_markup)


@dp.callback_query_handler(is_not_block_filter, text="calories_count")
async def calories_count(callmes: types.CallbackQuery | types.Message, state: FSMContext):
    db.register_activity(callmes.from_user.id, "calories_count", datetime.now(), date.today())
    if isinstance(callmes, types.CallbackQuery):
        logging.info(f"{sys._getframe().f_code.co_name} | {callmes.message.chat.id = }; {callmes.from_user.username = }; {callmes.data = }; {await state.get_data() = }")
        message = callmes.message
    else:
        logging.info(f"{sys._getframe().f_code.co_name} | {callmes.chat.id = }; {callmes.from_user.username = }; {await state.get_data() = }")
        message = callmes

    have_a_profile = db.have_user_a_profile(message.chat.id)

    if have_a_profile == 2:
        user_profile = db.get_user_profile(message.chat.id)

        if user_profile[2]:
            age = user_profile[2]
        else:
            age = date.today().year - datetime.strptime(user_profile[3], "%Y-%m-%d").year
            if date.today().replace(year=1) < datetime.strptime(user_profile[3], "%Y-%m-%d").date().replace(year=1):
                age -= 1

        if user_profile[4] == "male":
            bmr = (88.362 + (13.397 * user_profile[1]) + (4.799 * (user_profile[0] * 100)) - (5.677 * age))
        else:
            bmr = (447.593 + (9.247 * user_profile[1]) + (3.098 * (user_profile[0] * 100)) - (4.33 * age))

        norma = bmr * LEVEL_OF_ACTIVITIES_LIST[user_profile[-1]]['ratio']

        text = f"💤 Ваш <b>базовый обмен веществ</b>: {round(bmr)} ккал\n2️⃣4️⃣ Ваша <b>суточная норма</b>: {round(norma)} ккал\n🏃 Количество ккал <b>для похудения</b>: {round(norma * 0.8)} - {round(norma * 0.9)}\n💪 Количество ккал <b>для набора массы</b>: {round(norma * 1.1)} - {round(norma * 1.2)}\n\n❕ Данные рассчитаны от <i>ваших параметров</i>"

        await message.edit_text(text, reply_markup=to_menu_person_profile_markup)
    elif have_a_profile == 1:
        await CreateOrEditProfile.start(message, state, after_calories_count, 2)
    else:
        await CreateOrEditProfile.start(message, state, after_calories_count, 1)


async def after_calories_count(callmes: types.CallbackQuery | types.Message, state: FSMContext):
    if isinstance(callmes, types.CallbackQuery):
        logging.info(f"{sys._getframe().f_code.co_name} | {callmes.message.chat.id = }; {callmes.from_user.username = }; {callmes.data = }; {await state.get_data() = }")
        message = callmes.message
    else:
        logging.info(f"{sys._getframe().f_code.co_name} | {callmes.chat.id = }; {callmes.from_user.username = }; {await state.get_data() = }")
        message = callmes

    if db.have_user_a_profile(message.chat.id) == 2:
        await calories_count(callmes, state)
    else:
        await person_calories(callmes)


@dp.callback_query_handler(is_not_block_filter, text="imt_count")
async def imt_count(callmes: types.CallbackQuery | types.Message, state: FSMContext):
    db.register_activity(callmes.from_user.id, "imt_count", datetime.now(), date.today())
    if isinstance(callmes, types.CallbackQuery):
        logging.info(f"{sys._getframe().f_code.co_name} | {callmes.message.chat.id = }; {callmes.from_user.username = }; {callmes.data = }; {await state.get_data() = }")
        message = callmes.message
    else:
        logging.info(f"{sys._getframe().f_code.co_name} | {callmes.chat.id = }; {callmes.from_user.username = }; {await state.get_data() = }")
        message = callmes

    if db.have_user_a_profile(message.chat.id):
        user_profile = db.get_user_profile(message.chat.id)
        height = user_profile[0]
        weight = user_profile[1]

        imt = round(weight / height ** 2, 2)
        text = f"▫️ Ваш <b>ИМТ</b>: {imt}\n▫️ ИМТ <b>соответствует</b>: {[v for k, v in IMT_DICT.items() if k(imt)][0]}\n▫️ Ваш <b>идеальный вес</b>: {round(height ** 2 * 18.5, 1)} - {round(height ** 2 * 25, 1)}\n\n❕ Данные рассчитаны от <i>ваших параметров</i>"
        if isinstance(callmes, types.CallbackQuery):
            await message.edit_text(text, reply_markup=to_menu_person_profile_markup)
        else:
            await message.answer(text, reply_markup=to_menu_person_profile_markup)
    else:
        await CreateOrEditProfile.start(message, state, after_imt_count, 0)


async def after_imt_count(message: types.Message, state: FSMContext):
    logging.info(f"{sys._getframe().f_code.co_name} | {message.chat.id = }; {message.from_user.username = }; {await state.get_data() = }")
    
    if db.have_user_a_profile(message.chat.id):
        await imt_count(message, state)
    else:
        await person_calories(message)


@dp.callback_query_handler(is_not_block_filter, text="my_profile")
async def my_profile(callmes: types.CallbackQuery | types.Message, state: FSMContext):
    if isinstance(callmes, types.CallbackQuery):
        logging.info(f"{sys._getframe().f_code.co_name} | {callmes.message.chat.id = }; {callmes.from_user.username = }; {callmes.data = }; {await state.get_data() = }")
        message = callmes.message
    else:
        logging.info(f"{sys._getframe().f_code.co_name} | {callmes.chat.id = }; {callmes.from_user.username = }; {await state.get_data() = }")
        message = callmes

    if db.have_user_a_profile(message.chat.id) == 2:
        user_profile = db.get_user_profile(message.chat.id)

        text = f"Ваши данные:\n\n📈 <b>Рост</b> - {user_profile[0]} м\n⚖️ <b>Вес</b> - {user_profile[1]} кг\n🚹 <b>Пол</b> - {GENDER_DICT[user_profile[4]]}\n🏃 <b>Уровень активности</b> - {LEVEL_OF_ACTIVITIES_LIST[user_profile[5]]['title']}"

        if user_profile[2]:
            text += f"\n🔟 <b>Возраст</b> - {user_profile[2]}"
        else:
            age = date.today().year - datetime.strptime(user_profile[3], "%Y-%m-%d").year
            if date.today().replace(year=1) < datetime.strptime(user_profile[3], "%Y-%m-%d").date().replace(year=1):
                age -= 1

            text += f"\n🔟 <b>Возраст</b> - {age}\n👶 <b>Дата рождения</b> - {datetime.strptime(user_profile[3], '%Y-%m-%d').strftime('%d.%m.%Y')}"

        if isinstance(callmes, types.CallbackQuery):
            await message.edit_text(text, reply_markup=edit_user_profile_markup)
        else:
            await message.answer(text, reply_markup=edit_user_profile_markup)
    elif db.have_user_a_profile(message.chat.id) == 1:
        user_profile = db.get_user_profile(message.chat.id)

        text = f"Ваши данные:\n\n📈 <b>Рост</b> - {user_profile[0]} м\n⚖️ <b>Вес</b> - {user_profile[1]} кг\n🚹 <b>Пол</b> -\n🏃 <b>Уровень активности</b> -\n🔟 <b>Возраст</b> -\n👶 <b>Дата рождения</b> -\n\nНажмите на кнопку чтобы добавить <i>оставшиеся данные</i>"

        if isinstance(callmes, types.CallbackQuery):
            await message.edit_text(text, reply_markup=finish_create_user_profile_markup)
        else:
            await message.answer(text, reply_markup=finish_create_user_profile_markup)
    else:
        if isinstance(callmes, types.CallbackQuery):
            await message.edit_text(f'✖️ Вы еще не добавили <i>данные</i>.\n\nНажмите на кнопку чтобы их добавить',
                                    reply_markup=create_user_profile_markup)
        else:
            await message.answer(f'✖️ Вы еще не добавили <i>данные</i>.\n\nНажмите на кнопку чтобы их добавить',
                                 reply_markup=create_user_profile_markup)


@dp.callback_query_handler(is_not_block_filter, text="create_or_edit_profile")
async def create_or_edit_profile(callback: types.CallbackQuery, state: FSMContext):
    logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {await state.get_data() = }")
    
    await CreateOrEditProfile.start(callback.message, state, my_profile, 1)


@dp.callback_query_handler(is_not_block_filter, text="finish_create_profile")
async def finish_create_profile(callback: types.CallbackQuery, state: FSMContext):
    logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {await state.get_data() = }")
    
    await CreateOrEditProfile.start(callback.message, state, my_profile, 2)


@dp.callback_query_handler(is_not_block_filter, text="edit_level_of_activities")
async def edit_level_of_activities(callback: types.CallbackQuery, state: FSMContext):
    logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {await state.get_data() = }")
    
    await CreateOrEditProfile.start(callback.message, state, my_profile, 3)


class CreateOrEditProfile:
    @staticmethod
    async def cancel(message: types.Message, state: FSMContext):
        logging.debug(f"CreateOrEditProfile.cancel | {message.chat.id = }; {message.from_user.username = }; {await state.get_state() = }; {await state.get_data() = }")
        if message.text == "✖️ Отмена":
            logging.info(f"CreateOrEditProfile.cancel | Пользователь нажал \"Отмена\"; {message.chat.id = }; {message.from_user.username = }; {await state.get_state() = }; {await state.get_data() = }")
            data = await state.get_data()
            await state.finish()

            message = await message.answer(text="🕐 Загрузка ...", reply_markup=remove_markup)
            await message.delete()

            await data['func'](message, state)
            await send_mailing(message.chat.id)

            return False
        return True

    @staticmethod
    async def start(message: types.Message, state: FSMContext, func=None, start=None):
        logging.info(f"CreateOrEditProfile.start | {message.chat.id = }; {message.from_user.username = }; {func = }; {start = }; {await state.get_state() = }; {await state.get_data() = }")
        end_message = db.get_end_message(message.chat.id)

        if func and start in [0, 1]:
            await CreateOrEditProfileStatesGroup.first()

            await message.delete()
            await message.answer("📈 Отправьте свой <b>рост</b> (см) в виде числа:",
                                 reply_markup=cancel_markup)
            db.update_end_message(0, message.chat.id)
        elif func and start == 2:
            await CreateOrEditProfileStatesGroup.gender.set()
            if end_message == 0:
                await message.edit_text("🚹 Выберите <b>пол</b>:", reply_markup=choose_the_gender_markup)
            else:
                await message.delete()
                await message.answer("🚹 Выберите <b>пол</b>:", reply_markup=choose_the_gender_markup)
                db.update_end_message(0, message.chat.id)
        elif func and start == 3:
            await CreateOrEditProfileStatesGroup.level_of_activities.set()
            if end_message == 0:
                await message.edit_text("🏃 Выберите свой <b>уровень активности</b>:", reply_markup=choose_level_of_activities_markup)
            else:
                await message.delete()
                await message.answer("🏃 Выберите свой <b>уровень активности</b>:", reply_markup=choose_level_of_activities_markup)
                db.update_end_message(0, message.chat.id)
        else:
            raise TypeError("Укажите аргумент func")

        await state.update_data(start=start, func=func)

    @staticmethod
    @dp.message_handler(is_not_block_filter, content_types=['text'], state=CreateOrEditProfileStatesGroup.height)
    async def get_height(message: types.Message, state: FSMContext):
        logging.info(f"CreateOrEditProfile.{sys._getframe().f_code.co_name} | {message.chat.id = }; {message.from_user.username = }; {await state.get_state() = }; {await state.get_data() = }")
        if await CreateOrEditProfile.cancel(message, state):
            try:
                height = float(message.text.replace(",", "."))
                if not 10 <= height <= 300:
                    raise Exception

                await state.update_data(height=height / 100)
                await CreateOrEditProfileStatesGroup.next()
                await message.answer("⚖️ Отправьте свой <b>вес</b> (кг) в виде числа:", reply_markup=cancel_markup)
            except Exception as e:
                logging.warning(f"get_height | Error - {e}; {e.args = }")
                await message.answer(
                    "❌ <i>Неверный формат</i>.\n\n📈 Отправьте свой <b>рост</b> (см) в виде числа ещё раз:",
                    reply_markup=cancel_markup)

    @staticmethod
    @dp.message_handler(is_not_block_filter, content_types=['text'], state=CreateOrEditProfileStatesGroup.weight)
    async def get_weight(message: types.Message, state: FSMContext):
        logging.info(f"CreateOrEditProfile.{sys._getframe().f_code.co_name} | {message.chat.id = }; {message.from_user.username = }; {await state.get_state() = }; {await state.get_data() = }")
        if await CreateOrEditProfile.cancel(message, state):
            try:
                weight = float(message.text.replace(",", ('.')))
                if not 10 <= weight <= 400:
                    raise Exception

                message = await message.answer(text="🕐 Загрузка ...", reply_markup=remove_markup)
                await message.delete()

                data = await state.get_data()
                start = data['start']

                if start == 0:
                    height = data['height']

                    await state.finish()

                    db.add_user_weight_and_height(message.chat.id, round(height, 3), round(weight, 2))

                    await data['func'](message, state)
                    await send_mailing(message.chat.id)
                else:
                    await state.update_data(weight=weight)
                    await CreateOrEditProfileStatesGroup.next()
                    await message.answer("🚹 Выберите <b>пол</b>:", reply_markup=choose_the_gender_markup)
            except Exception as e:
                logging.warning(f"get_weight | Error - {e}; {e.args = }")
                await message.answer(
                    "❌ <i>Неверный формат</i>.\n\n️Отправьте свой <b>вес</b> (кг) в виде числа ещё раз:",
                    reply_markup=cancel_markup)

    @staticmethod
    @dp.callback_query_handler(is_not_block_filter, state=CreateOrEditProfileStatesGroup.gender)
    async def get_gender(callback: types.CallbackQuery, state: FSMContext):
        logging.info(f"CreateOrEditProfile.{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {await state.get_state() = }; {await state.get_data() = }")
        if callback.data == "cancel":
            data = await state.get_data()
            await state.finish()
            await data['func'](callback, state)
            await send_mailing(callback.message.chat.id)
        else:
            await state.update_data(gender=callback.data)
            await CreateOrEditProfileStatesGroup.next()

            await callback.message.delete()
            await callback.message.answer(
                "Отправьте свой <b>🔟 возраст</b> в виде целого числа (25) или <b>👶 дату рождения</b> в следующем формате - 14-06-1995:",
                reply_markup=cancel_markup)

    @staticmethod
    @dp.callback_query_handler(is_not_block_filter, text="to_choose_level_of_activities", state="*")
    @dp.message_handler(is_not_block_filter, content_types=['text'], state=CreateOrEditProfileStatesGroup.age)
    async def get_age(mescal: types.Message | types.CallbackQuery, state: FSMContext):
        if isinstance(mescal, types.Message):
            logging.info(f"CreateOrEditProfile.{sys._getframe().f_code.co_name} | {mescal.chat.id = }; {mescal.from_user.username = }; {await state.get_state() = }; {await state.get_data() = }")
            text = mescal.text
            age = None
            born_date = None

            if await CreateOrEditProfile.cancel(mescal, state):
                try:
                    if "-" in text:
                        born_date = datetime.strptime(text, "%d-%m-%Y").date()
                        today = date.today()

                        if not date(year=today.year - 150, month=today.month, day=today.day) <= born_date <= date(
                                year=today.year - 1, month=today.month, day=today.day):
                            raise Exception
                    else:
                        age = int(text)

                        if not 1 <= age <= 150:
                            raise Exception

                    message = await mescal.answer(text="🕐 Загрузка ...", reply_markup=remove_markup)
                    await message.delete()

                    await state.update_data(age=age, born_date=born_date)
                    await CreateOrEditProfileStatesGroup.next()
                    await mescal.answer("🏃 Выберите свой <b>уровень активности</b>:",
                                        reply_markup=choose_level_of_activities_markup)
                except Exception as e:
                    logging.warning(f"get_age | Error - {e}; {e.args = }")
                    await mescal.answer(
                        "❌ <i>Неверный формат</i>.\n\n️Отправьте свой <b>🔟 возраст</b> в виде целого числа (25) или <b>👶 дату рождения</b> в следующем формате - 14-06-1995 ещё раз:",
                        reply_markup=cancel_markup)
        else:
            logging.info(f"CreateOrEditProfile.{sys._getframe().f_code.co_name} | {mescal.message.chat.id = }; {mescal.from_user.username = }; {mescal.data = }; {await state.get_state() = }; {await state.get_data() = }")
            await mescal.message.edit_text("🏃 Выберите свой <b>уровень активности</b>:",
                                           reply_markup=choose_level_of_activities_markup)

    @staticmethod
    @dp.callback_query_handler(is_not_block_filter, text="how_choose_level_of_activities", state="*")
    async def how_choose_level_of_activities(callback: types.CallbackQuery, state: FSMContext):
        logging.info(f"CreateOrEditProfile.{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {await state.get_state() = }; {await state.get_data() = }")
        await callback.message.edit_text(
            "Вам нужно выбрать один из пяти вариантов <i>нагрузки</i> — от <b>минимального</b> до <b>очень высокого</b>. Здесь учитываются не только упражнения, но и <i>прочая активность</i>: <b>пешие прогулки</b>, <b>работа по дому</b>. Чтобы полученные цифры соответствовали <i>действительности</i>, важно <b>не преувеличивать</b> <i>характеристики</i>.\n\nВыбирайте <b>минимальную нагрузку</b>, если работаете в офисе, не тренируетесь, а вечера и выходные предпочитаете проводить за компьютером, а не на прогулках.\n\nПараметр «<b>низкая нагрузка</b>» подойдёт для тех, кто много трудится по дому, гуляет с собакой, иногда выбирает прогулку вместо транспорта и изредка занимается спортом.\n\n<b>Средняя нагрузка</b> подразумевает, что вы тренируетесь от 3 до 5 раз в неделю, при этом в свободное время не только лежите на диване, но и ходите пешком, выполняете бытовые задачи.\n\n<b>Высокая нагрузка</b> предполагает, что вы занимаетесь спортом 6–7 раз в неделю — или 3–5, но при этом ваша работа — физический труд.\n\n<b>Очень высокая нагрузка</b> характерна для тех, кто тренируется по 2 раза в день или много занимается спортом и работает физически, но при этом и о других видах активности не забывает.",
            reply_markup=to_choose_level_of_activities_markup)

    @staticmethod
    @dp.callback_query_handler(is_not_block_filter, state=CreateOrEditProfileStatesGroup.level_of_activities)
    async def get_level_of_activities(callback: types.CallbackQuery, state: FSMContext):
        data = await state.get_data()
        logging.info(f"CreateOrEditProfile.{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {await state.get_state() = }; {data = }")
        await state.finish()

        if callback.data != "cancel":
            if data['start'] == 1:
                db.add_user_profile(callback.message.chat.id, data['height'], data['weight'], data['age'],
                                    data['born_date'], data['gender'], int(callback.data))
            elif data['start'] == 2:
                db.add_end_of_user_profile(callback.message.chat.id, data['age'], data['born_date'], data['gender'],
                                           int(callback.data))
            elif data['start'] == 3:
                db.add_level_of_activities(callback.message.chat.id, int(callback.data))

        await data['func'](callback, state)
        await send_mailing(callback.message.chat.id)


@dp.callback_query_handler(is_not_block_filter, text="stats")
async def stats(callback: types.CallbackQuery):
    logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }")

    await callback.message.edit_text(text="Выберите <b>пункт</b>:", reply_markup=stats_markup)


@dp.callback_query_handler(is_not_block_filter, text="stats_photo")
async def to_menu_photo(callback: types.CallbackQuery):
    logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }")

    await callback.message.delete()
    await callback.message.answer(text="Выберите <b>пункт</b>:", reply_markup=stats_markup)
    db.update_end_message(0, callback.message.chat.id)

@dp.callback_query_handler(is_not_block_filter, text="count_of_users")
async def count_of_users(callback: types.CallbackQuery):
    logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }")

    if (date.today() - bot_create_date).days > 152:
        if bot_create_date.year == date.today().year:
            year = date.today().year
            x_list = [f"{bot_create_date.day}.{bot_create_date.month}.{bot_create_date.year}"] + [f"1.{month}.{year}"
                                                                                                  for month in range(
                    bot_create_date.month + 1, date.today().month, 3)] + [
                         f"{date.today().day}.{date.today().month}.{date.today().year}"]

        else:
            if bot_create_date.year + 1 == date.today().year:
                x_list = [f"{bot_create_date.day}.{bot_create_date.month}.{bot_create_date.year}"] + [
                    f"1.1.{date.today().year}"] + [
                             f"{date.today().day}.{date.today().month}.{date.today().year}"]
            else:
                x_list = [f"{bot_create_date.day}.{bot_create_date.month}.{bot_create_date.year}"] + [f"1.1.{year}" for
                                                                                                      year in range(
                        bot_create_date.year + 1, date.today().year,
                        math.ceil((date.today().year - bot_create_date.year + 1) / 4))] + [
                             f"{date.today().day}.{date.today().month}.{date.today().year}"]
    else:
        if bot_create_date.year == date.today().year:
            year = date.today().year
            x_list = [f"{bot_create_date.day}.{bot_create_date.month}.{bot_create_date.year}"] + [f"1.{month}.{year}"
                                                                                                  for month in range(
                    bot_create_date.month + 1, date.today().month)] + [
                         f"{date.today().day}.{date.today().month}.{date.today().year}"]
        else:
            x_list = [f"{bot_create_date.day}.{bot_create_date.month}.{bot_create_date.year}"] + [
                f"1.{month}.{bot_create_date.year}" for month in range(bot_create_date.month + 1, 13)] + [
                         f"1.{month}.{date.today().year}" for month in range(1, date.today().month)] + [
                         f"{date.today().day}.{date.today().month}.{date.today().year}"]

    x_list = [datetime.strptime(date_, "%d.%m.%Y").strftime('%d.%m.%Y') for date_ in x_list]
    y_list = [db.get_count_of_users_before_date(datetime.strptime(date_, "%d.%m.%Y").date()) for date_ in x_list]

    count_of_users = db.get_count_of_users()

    plt.title("Кол-во пользователей")
    plt.xlabel("Дата")
    plt.ylabel("Количество")
    plt.yticks((list(range(0, count_of_users + 1, math.ceil(count_of_users / 10)))) if (
                count_of_users % math.ceil(count_of_users / 10) == 0) else (
                list(range(0, count_of_users + 1, math.ceil(count_of_users / 10))) + [count_of_users]))
    if len(y_list) == 1:
        plt.bar(x_list, y_list)
    else:
        plt.plot(list(filter(bool, x_list)), y_list, marker="o")
    plt.savefig(os.path.join(path_to_data_dir, "count_of_users.png"))
    plt.close()

    await callback.message.delete()
    with open(os.path.join(path_to_data_dir, "count_of_users.png"), "rb") as file:
        await callback.message.answer_photo(photo=file,
                                            caption=f"🤖 <b>Версия бота</b> - {bot_version}\n👶 <b>Дата создания бота</b> - {bot_create_date.strftime('%d.%m.%Y')}\n👥 <b>Всего пользователей</b> - {count_of_users}",
                                            reply_markup=count_of_users_markup)
    db.update_end_message(0, callback.message.chat.id)


@dp.callback_query_handler(is_not_block_filter, text="main_stats")
async def main_stats(callback: types.CallbackQuery):
    logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }")

    tomorrow = date.today() + timedelta(days=1)
    update_timedelta = (datetime(tomorrow.year, tomorrow.month, tomorrow.day) - datetime.now())

    text = f"🤖 <b>О боте</b>\nВерсия бота - {bot_version}\nДата создания бота - {bot_create_date.strftime('%d.%m.%Y')}\n\n"
    text += f"🕐 <b>Статистика за все время</b>\nВсего пользователей - {db.get_count_of_users()}\nКоличество пользователей, которые сохранили данные о себе - {db.get_count_of_users_who_have_a_profile()}\nКоличество пользователей, которые смотрели калорийность продуктов - {db.get_count_from_user_activities(['item_products', 'item_answer_products_search'])}\nКоличество пользователей, которые смотрели рецепты - {db.get_count_from_user_activities(['item_recipes', 'item_answer_recipes_search'])}\nКоличество пользователей, которые рассчитывали суточную норму калорий  - {db.get_count_from_user_activities(['calories_count'])}\nКоличество пользователей, которые расчитывали свой идеальный вес и ИМТ - {db.get_count_from_user_activities(['imt_count'])}\n\n"
    text += f"2️⃣4️⃣ <b>Статистика за сегодня - {date.today().strftime('%d.%m.%Y')}. Новый день через {int(update_timedelta.total_seconds() // 3600)} час(а,ов) и {int((update_timedelta.total_seconds() % 3600) // 60)} минут(ы)</b>\nНовых пользователей - {db.get_count_of_users_in_date(date.today())}\nКоличество пользователей, которые смотрели калорийность продуктов - {db.get_count_from_user_activities_in_date(['item_products', 'item_answer_products_search'], date.today())}\nКоличество пользователей, которые смотрели рецепты - {db.get_count_from_user_activities_in_date(['item_recipes', 'item_answer_recipes_search'], date.today())}\nКоличество пользователей, которые рассчитывали суточную норму калорий  - {db.get_count_from_user_activities_in_date(['calories_count'], date.today())}\nКоличество пользователей, которые расчитывали свой идеальный вес и ИМТ - {db.get_count_from_user_activities_in_date(['imt_count'], date.today())}"

    await callback.message.edit_text(text=text, reply_markup=main_stats_markup)


@dp.callback_query_handler(is_not_block_filter, text="about_us")
async def about_us(callback: types.CallbackQuery):
    logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }")
    text = f"👨‍💻 <b>Программист</b>, который написал этого бота - {programmer}\n\n"
    if bot_owner == ad:
        text += f"👤 <b>Владелец бота</b> - {bot_owner}. <b>По рекламе</b> и всем другим вопросам обращаться к нему.\n"
    else:
        text += f"👤 <b>Владелец бота</b> - {bot_owner}\n💲 <b>По рекламе</b> и всем другим вопросам обращаться сюда - {ad}\n"

    if customer_service == bot_owner:
        text += f"😵 Если найдете <b>баг</b>, то просьба написать об этом владельцу бота"
    else:
        text += f"😵 Если найдете <b>баг</b>, то просьба написать об сюда - {customer_service}"

    await callback.message.edit_text(text=text, reply_markup=to_menu_markup)


@dp.callback_query_handler(is_not_block_filter, text="menu")
async def to_menu(callback: types.CallbackQuery, state: FSMContext):
    logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {await state.get_data() = }")
    
    await state.reset_data()

    await callback.message.edit_text(
        f"Этот бот поможет найти <b>калорийность продуктов</b> и <b>крутые рецепты</b>, а также вычислить <b>дневную норму калорий</b> для тебя.\nТакже ты сможешь учитывать <b>количество съеденных калорий</b>",
        reply_markup=menu_markup)


@dp.callback_query_handler(is_not_block_filter, text="menu_photo")
async def to_menu_photo(callback: types.CallbackQuery):
    logging.info(f"{sys._getframe().f_code.co_name} | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }")

    await callback.message.delete()
    await callback.message.answer(
        f"Этот бот поможет найти <b>калорийность продуктов</b> и <b>крутые рецепты</b>, а также вычислить <b>дневную норму калорий</b> для тебя.\nТакже ты сможешь учитывать <b>количество съеденных калорий</b>",
        reply_markup=menu_markup)
    db.update_end_message(0, callback.message.chat.id)

@dp.message_handler(is_admin_filter, commands=['admin_panel'])
@dp.callback_query_handler(is_admin_filter, text="admin_panel")
async def admin_panel(callmes: types.Message | types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    if isinstance(callmes, types.Message):
        logging.info(f"{sys._getframe().f_code.co_name} | Админ; {callmes.chat.id = }; {callmes.from_user.username}; {data = }")
    else:
        logging.info(f"{sys._getframe().f_code.co_name} | Админ; {callmes.message.chat.id = }; {callmes.from_user.username}; {callmes.data = }; {data = }")

    if "item_last_mailings" in data:
            if "previous_message_id" in data['item_last_mailings']:
                if isinstance(callmes, types.CallbackQuery):
                    await bot.delete_message(callmes.message.chat.id, data['item_last_mailings']['previous_message_id'])

    await state.reset_data()

    if isinstance(callmes, types.Message):
        await callmes.answer("Нажмите <b>на кнопку</b>:", reply_markup=admin_panel_markup)
        db.update_end_message(0, callmes.chat.id)
    else:
        await callmes.message.edit_text("Нажмите <b>на кнопку</b>:", reply_markup=admin_panel_markup)

@dp.callback_query_handler(is_admin_filter, text="mailings")
async def mailings(callback: types.CallbackQuery, state: FSMContext):
    logging.info(f"{sys._getframe().f_code.co_name} | Админ; {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {await state.get_data() = }")
    await state.reset_data()

    await callback.message.edit_text(text="Нажмите <b>на кнопку</b>:", reply_markup=mailings_markup_with_last if db.get_last_mailings() else mailings_markup)

@dp.callback_query_handler(is_admin_filter, text="create_mailing")
async def create_mailing(callback: types.CallbackQuery):
    logging.info(f"{sys._getframe().f_code.co_name} | Админ; {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }")
    await Mailing.start(callback)

class Mailing:
    @staticmethod
    async def cancel(message: types.Message, state: FSMContext):
        logging.debug(f"Mailing.cancel | Админ; {message.chat.id = }; {message.from_user.username = }; {await state.get_state() = }; {await state.get_data() = }")
        if message.text == "✖️ Отмена":
            logging.info(f"Mailing.cancel | Админ; Пользователь нажал \"Отмена\"; {message.chat.id = }; {message.from_user.username = }; {await state.get_state() = }; {await state.get_data() = }")
            await state.finish()

            await message.answer(text="Нажмите <b>на кнопку</b>:", reply_markup=mailings_markup_with_last if db.get_last_mailings() else mailings_markup)
            await send_mailing(message.chat.id)
            return False
        return True

    @staticmethod
    async def start(callback: types.CallbackQuery):
        logging.info(f"Mailing.start | Админ; {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }")
        await MailingStatesGroup.first()

        await callback.message.delete()
        await callback.message.answer("Отправьте сообщение для рассылки:", reply_markup=cancel_markup)
        db.update_end_message(0, callback.message.chat.id)

    @staticmethod
    @dp.message_handler(is_admin_filter, content_types=['text'], state=MailingStatesGroup.get_message)
    async def get_message(message: types.Message, state: FSMContext):
        logging.info(f"{sys._getframe().f_code.co_name} | Админ; {message.chat.id = }; {message.from_user.username = }; {await state.get_state() = }; {await state.get_data() = }")
        messaged = await message.answer(text="🕐 Загрузка ...", reply_markup=remove_markup)
        await messaged.delete()

        if await Mailing.cancel(message, state):
            await state.update_data(message_for_mailing=message)
            await MailingStatesGroup.next()

            await message.answer("Начать рассылку?", reply_markup=confirmation_markup)

    @staticmethod
    @dp.callback_query_handler(is_admin_filter, state=MailingStatesGroup.get_confirmation)
    async def get_confirmation(callback: types.CallbackQuery, state: FSMContext):
        data = await state.get_data()
        logging.info(f"{sys._getframe().f_code.co_name} | Админ; {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {await state.get_state() = }; {data = }")
        message_for_mailng = data["message_for_mailing"]

        await state.finish()
        
        now_datetime = datetime.now()
        all_count = db.get_count_of_users_without_admins(admins)
        if callback.data == "no":
            db.add_mailing(message_for_mailng.message_id, message_for_mailng.chat.id, all_count, now_datetime, 0)
            await callback.message.edit_text("❌ <i>Рассылка отменена</i>", reply_markup=mailings_markup_with_last)
            await send_mailing(callback.message.chat.id)
        elif callback.data == "yes":
            db.add_mailing(message_for_mailng.message_id, message_for_mailng.chat.id, all_count, now_datetime, 1)
            await callback.message.edit_text("✅ <i>Рассылка начата</i>", reply_markup=mailings_markup_with_last)
            await send_mailing(callback.message.chat.id)

            success_count = 0
            unsuccess_count = 0
            later_count = 0

            for chat_id in db.get_all_chat_ids():
                if db.get_type_of_mailing_by_message_id(message_for_mailng.message_id) == 0:
                    return
                if db.is_user_had_sent_notion_without_mailing_id(message_for_mailng.message_id, chat_id[0]):
                    success_count += 1
                elif int(chat_id[0]) not in admins:
                    user_state = dp.current_state(chat=chat_id[0], user=chat_id[0])
                    if await user_state.get_state():
                        later_count += 1
                    else:
                        try:
                            await message_for_mailng.copy_to(chat_id=chat_id[0])
                        except:
                            logging.warning(f"get_confirmation | Не удалось отправить рассылку - {chat_id[0]}")
                            unsuccess_count += 1
                        else:
                            logging.info(f"get_confirmation | Удалось отправить рассылку - {chat_id[0]}")
                            db.update_success_count_in_mailing(message_for_mailng.message_id)
                            db.add_mailings_users(message_for_mailng.message_id, chat_id[0])
                            db.update_end_message(1, chat_id[0])
                            success_count += 1

            if not await state.get_state():
                db.update_end_message(1, callback.message.chat.id)
                await callback.message.answer(f"<i>Рассылка созданная в {now_datetime.strftime('%d.%m.%Y %H:%M:%S')} завершена</i>\n\n<b>Всего</b> - {all_count}\n<b>Успешно</b> - {success_count}\n<b>Неуспешно</b> - {unsuccess_count}\n<b>Будет отправлено позже</b> - {later_count}")
            else:
                db.add_end_of_mailing(f"<i>Рассылка созданная в {now_datetime.strftime('%d.%m.%Y %H:%M:%S')} завершена</i>\n\n<b>Всего</b> - {all_count}\n<b>Успешно</b> - {success_count}\n<b>Неуспешно</b> - {unsuccess_count}\n<b>Будет отправлено позже</b> - {later_count}", callback.message.chat.id)

@dp.callback_query_handler(is_admin_filter, lambda call: call.data.startswith("lastmailings"))
async def last_mailings(callback: types.CallbackQuery, state: FSMContext):
    async with state.proxy() as data:
        logging.info(f"{sys._getframe().f_code.co_name} | Админ; {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {data.as_dict() = }")
        if "last_mailings" in data:
            func_data = data["last_mailings"]
        else:
            func_data = {}

        if "paginator" in func_data:
            paginator = func_data["paginator"]
        else:
            paginator = Paginator(data=list(map(lambda x: [x[0], datetime.fromisoformat(x[1]).strftime('%d.%m.%Y %H:%M:%S'), *x[2:]], db.get_last_mailings())),
                                  callback_prefix=f"lastmailings_0",
                                  back_callback=f"mailings",
                                  type="itemlist", width=2, height=4, item_prefix="itemlastmailings",
                                  category_id=0, second_type="mailings")
            func_data["paginator"] = paginator

        func_data['page'] = int(callback.data.split("_")[-1])
        data['last_mailings'] = func_data

        if "item_last_mailings" in data:
            if "previous_message_id" in data['item_last_mailings']:
                await bot.delete_message(callback.message.chat.id, data['item_last_mailings']['previous_message_id'])
                del data['item_last_mailings']['previous_message_id']

    await callback.message.edit_text(text="Выберите <b>рассылку</b>",
                                     reply_markup=paginator.get_page_keyboard(callback.data, [i for i in range(len(paginator.get_data()))]))


@dp.callback_query_handler(is_admin_filter, lambda call: call.data.startswith("itemlastmailings"))
async def item_last_mailings(callback: types.CallbackQuery, state: FSMContext):
    async with state.proxy() as data:
        logging.info(f"{sys._getframe().f_code.co_name} | Админ; {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {data.as_dict() = }")
        if "item_last_mailings" in data:
            func_data = data[f'item_last_mailings']
        else:
            func_data = {}

        page_last_mailings = data["last_mailings"]["page"]
        if "paginator" in func_data:
            paginator = func_data['paginator']

            n_page_last_mailings = data["last_mailings"]["paginator"].get_page_by_data_ind(int(callback.data.split('_')[-1]))
            if page_last_mailings != n_page_last_mailings:
                data["last_mailings"]["page"] = n_page_last_mailings
                paginator.back_callback = f"lastmailings_0_{n_page_last_mailings}"
            elif int(paginator.back_callback.split("_")[-1]) != n_page_last_mailings:
                paginator.back_callback = f"lastmailings_0_{n_page_last_mailings}"
        else:
            paginator = Paginator(data=data['last_mailings']['paginator'].get_data(),
                                  callback_prefix=f"itemlastmailings_0",
                                  back_callback=f"lastmailings_0_{page_last_mailings}",
                                  type="items", width=2, second_type="mailings")
            func_data['paginator'] = paginator

    mailing = paginator.get_data()[int(callback.data.split("_")[-1])]
    type_of_mailing = db.get_type_of_mailing_by_mailing_id(mailing[0])
    success_count_in_mailing = db.get_success_count_in_mailing(mailing[0])

    if "previous_message_id" in func_data:
        await bot.delete_message(callback.message.chat.id, func_data["previous_message_id"])
    await callback.message.delete()
    copy_message = await bot.copy_message(callback.message.chat.id, mailing[3], mailing[2])
    func_data['previous_message_id'] = copy_message.message_id
    await state.update_data(item_last_mailings=func_data)
    await callback.message.answer(text=f"<b>Дата создания рассылки</b> - {mailing[1]}\n<b>Всего</b> - {mailing[4]}\n<b>Успешно</b> - {success_count_in_mailing}\n<b>Статус</b> - {'Отменена' if type_of_mailing == 0 else ('Завершена' if success_count_in_mailing == mailing[4] else 'В работе')}", reply_markup=paginator.get_page_keyboard(callback.data, option=2 if success_count_in_mailing == mailing[4] else type_of_mailing))
    db.update_end_message(0, callback.message.chat.id)

@dp.callback_query_handler(is_admin_filter, lambda call: call.data.startswith("restartmailing"))
async def restart_mailing(callback: types.CallbackQuery, state: FSMContext):
    async with state.proxy() as data:
        logging.info(f"{sys._getframe().f_code.co_name} | Админ; {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {data.as_dict() = }")
        paginator = data['item_last_mailings']['paginator']
    mailing = paginator.get_data()[int(callback.data.split("_")[-1])]

    if db.get_type_of_mailing_by_mailing_id(mailing[0]) == 0:
        db.update_type_in_mailings(mailing[0], 1)

        await callback.message.edit_text(text=f"<b>Дата создания рассылки</b> - {mailing[1]}\n<b>Всего</b> - {mailing[4]}\n<b>Успешно</b> - {db.get_success_count_in_mailing(mailing[0])}\n<b>Статус</b> - В работе", reply_markup=paginator.get_page_keyboard(callback.data, option=1))

        for chat_id in db.get_not_is_sent_users(mailing[0]):
            if db.get_type_of_mailing_by_mailing_id(mailing[0]) == 0:
                return
            if int(chat_id[0]) not in admins and not db.is_user_had_sent_notion(mailing[0], chat_id[0]):
                user_state = dp.current_state(chat=chat_id[0], user=chat_id[0])
                if not await user_state.get_state():
                    try:
                        await bot.copy_message(chat_id[0], mailing[3], mailing[2])
                    except:
                        logging.warning(f"restart_mailing | Не удалось отправить рассылку - {chat_id[0]}")
                    else:
                        logging.info(f"restart_mailing | Удалось отправить рассылку - {chat_id[0]}")
                        db.add_mailings_users_with_mailing_id(mailing[0], chat_id[0])
                        db.update_success_count_in_mailing(mailing[2])
                        db.update_end_message(1, chat_id[0])
    else:
        success_count_in_mailing = db.get_success_count_in_mailing(mailing[0])

        await callback.message.edit_text(text=f"<b>Дата создания рассылки</b> - {mailing[1]}\n<b>Всего</b> - {mailing[4]}\n<b>Успешно</b> - {success_count_in_mailing}\n<b>Статус</b> - {'Завершена' if success_count_in_mailing == mailing[4] else 'В работе'}", reply_markup=paginator.get_page_keyboard(callback.data, option=2 if success_count_in_mailing == mailing[4] else 1))


@dp.callback_query_handler(is_admin_filter, lambda call: call.data.startswith("cancelmailing"))
async def cancel_mailing(callback: types.CallbackQuery, state: FSMContext):
    async with state.proxy() as data:
        logging.info(f"{sys._getframe().f_code.co_name} | Админ; {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {data.as_dict() = }")
        paginator = data['item_last_mailings']['paginator']
    mailing = paginator.get_data()[int(callback.data.split("_")[-1])]

    if db.get_type_of_mailing_by_mailing_id(mailing[0]) == 1:
        db.update_type_in_mailings(mailing[0], 0)

    await callback.message.edit_text(text=f"<b>Дата создания рассылки</b> - {mailing[1]}\n<b>Всего</b> - {mailing[4]}\n<b>Успешно</b> - {db.get_success_count_in_mailing(mailing[0])}\n<b>Статус</b> - Отменена", reply_markup=paginator.get_page_keyboard(callback.data, option=0))

@dp.callback_query_handler(is_admin_filter, text="users")
async def users_menu(callback: types.CallbackQuery, state: FSMContext):
    logging.info(f"{sys._getframe().f_code.co_name} | Админ; {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {await state.get_data() = }")
    await state.reset_data()

    await callback.message.edit_text(text="Выберите <b>пункт</b>:", reply_markup=users_markup)

@dp.callback_query_handler(is_admin_filter, text="userslist")
async def users_list(callback: types.CallbackQuery):
    logging.info(f"{sys._getframe().f_code.co_name} | Админ; {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }")
    await callback.message.edit_text(text="Выберите <b>формат файла</b>:", reply_markup=users_list_markup)

@dp.callback_query_handler(is_admin_filter, lambda call: call.data.startswith("userslist"))
async def user_list_extension(callback: types.CallbackQuery):
    logging.info(f"{sys._getframe().f_code.co_name} | Админ; {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }")

    extension = callback.data.split("_")[-1]
    users = db.get_users()
    path = os.path.join(path_to_data_dir, f"users.{extension}")
    
    if extension == "json":
        json_data = [dict(zip(JSON_KEYS, user)) for user in users]

        with open(path, "w", encoding="utf8") as file:
            json.dump({'data': json_data}, file, ensure_ascii=False, indent=4)
    elif extension == "xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active

        ws["A1"] = "id"
        ws["B1"] = "chat_id"
        ws["C1"] = "name"
        ws["D1"] = "username"
        ws["E1"] = "height"
        ws["F1"] = "weight"
        ws["G1"] = "age"
        ws["H1"] = "born_date"
        ws["I1"] = "gender"
        ws["J1"] = "level_of_activities"
        ws["K1"] = "end_message"
        ws["L1"] = "register_datetime"
        ws["M1"] = "register_date"

        for i, user in enumerate(users, 2):
            ws[f"A{i}"] = user[0]
            ws[f"B{i}"] = user[1]
            ws[f"C{i}"] = user[2]
            ws[f"D{i}"] = user[3]
            ws[f"E{i}"] = user[4]
            ws[f"F{i}"] = user[5]
            ws[f"G{i}"] = user[6]
            ws[f"H{i}"] = user[7]
            ws[f"I{i}"] = user[8]
            ws[f"J{i}"] = user[9]
            ws[f"K{i}"] = user[10]
            ws[f"L{i}"] = user[11]
            ws[f"M{i}"] = user[12]

        wb.save(path)
    else:
        if extension == "txt":
            file_text = TXT_START
            for user in users:
                file_text += "\n" + "; ".join(map(str, user))
        elif extension == "html":
            file_text = template.render(users=users)

        with open(path, "w", encoding="utf8") as file:
            file.write(file_text)

    db.update_end_message(1, callback.from_user.id)
    input_file = types.InputFile(path)
    await callback.message.answer_document(document=input_file, caption="Пользователи")

@dp.callback_query_handler(is_admin_filter, lambda call: call.data.startswith("blockusers"))
async def block_users(callback: types.CallbackQuery, state: FSMContext):
    logging.info(f"{sys._getframe().f_code.co_name} | Админ; {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {await state.get_data() = }")

    block_users_ = db.get_block_users()
    if not block_users_:
        await callback.answer("❗️ Вы ещё не блокировали ни одного пользователя", show_alert=True)
        return
    async with state.proxy() as data:
        if "block_users" in data:
            func_data = data["block_users"]
        else:
            func_data = {}

        if "paginator" in func_data:
            paginator = func_data["paginator"]
        else:
            paginator = Paginator(data=block_users_,
                                  callback_prefix=f"blockusers_0",
                                  back_callback=f"users",
                                  type="itemlist", width=2, height=4, item_prefix="itemblockusers",
                                  category_id=0, second_type="blockusers")
            func_data["paginator"] = paginator

        func_data['page'] = int(callback.data.split("_")[-1])
        data['block_users'] = func_data

    await callback.message.edit_text(text="Выберите <b>пользователя</b>",
                                     reply_markup=paginator.get_page_keyboard(callback.data, [i for i in range(len(paginator.get_data()))]))

@dp.callback_query_handler(is_admin_filter, lambda call: call.data.startswith("itemblockusers"))
async def item_block_users(callback: types.CallbackQuery, state: FSMContext):
    async with state.proxy() as data:
        logging.info(f"{sys._getframe().f_code.co_name} | Админ; {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {data.as_dict() = }")
        if "item_block_users" in data:
            func_data = data[f'item_block_users']
        else:
            func_data = {}

        page_block_users = data["block_users"]["page"]
        if "paginator" in func_data:
            paginator = func_data['paginator']

            n_page_block_users = data["block_users"]["paginator"].get_page_by_data_ind(int(callback.data.split('_')[-1]))
            if page_block_users != n_page_block_users:
                data["block_users"]["page"] = n_page_block_users
                paginator.back_callback = f"blockusers_0_{n_page_block_users}"
            elif int(paginator.back_callback.split("_")[-1]) != n_page_block_users:
                paginator.back_callback = f"blockusers_0_{n_page_block_users}"
        else:
            paginator = Paginator(data=data['block_users']['paginator'].get_data(),
                                  callback_prefix=f"itemblockusers_0",
                                  back_callback=f"blockusers_0_{page_block_users}",
                                  type="items", width=2, second_type="blockusers")
            func_data['paginator'] = paginator

        data["item_block_users"] = func_data

    user = paginator.get_data()[int(callback.data.split("_")[-1])]
    is_block = bool(db.is_user_block(user[0]))

    if user[3]:
        user_name, user_username = db.get_names_by_id(user[3])
        username_text = f"@{user_username}" if user_username else ''
        
        text = f"<b>ID</b> - {user[4]}\n<b>Username</b> - {username_text}\n<b>Имя</b> - {user_name}\n<b>Дата регистрации</b> - {datetime.fromisoformat(user[-1]).strftime('%d.%m.%Y %H:%M:%S')}"

        if is_block:
            text += f"\n<b>Дата блокировки</b> - {datetime.fromisoformat(db.get_block_datetime(user[0])).strftime('%d.%m.%Y %H:%M:%S')}\n\n<i>🙅‍♂️ Пользователь заблокирован</i>"
        else:
            text += f"\n<b>Дата разблокировки</b> - {datetime.fromisoformat(db.get_block_datetime(user[0])).strftime('%d.%m.%Y %H:%M:%S')}\n\n<i>🤵 Пользователь разблокирован</i>"

        await callback.message.edit_text(text=text, reply_markup=paginator.get_page_keyboard(callback.data, option=is_block))
    else:
        user_id = db.get_block_user_id(user[0])
        if user_id:
            _, user_chat_id, user_username, user_name, user_register_datetime = db.get_user_info_by_id(user_id)
            username_text = f"@{user_username}" if user_username else ''

            text = f"<b>ID</b> - {user_chat_id}\n<b>Username</b> - {username_text}\n<b>Имя</b> - {user_name}\n<b>Дата регистрации</b> - {datetime.fromisoformat(user_register_datetime).strftime('%d.%m.%Y %H:%M:%S')}"
            if is_block:
                text += f"\n<b>Дата блокировки</b> - {datetime.fromisoformat(db.get_block_datetime(user[0])).strftime('%d.%m.%Y %H:%M:%S')}\n\n<i>🙅‍♂️ Пользователь заблокирован</i>"
            else:
                text += f"\n<b>Дата разблокировки</b> - {datetime.fromisoformat(db.get_block_datetime(user[0])).strftime('%d.%m.%Y %H:%M:%S')}\n\n<i>🤵 Пользователь разблокирован</i>"

            await callback.message.edit_text(text=text, reply_markup=paginator.get_page_keyboard(callback.data, option=is_block))
        else:
            text = f"<b>Username</b> - {user[1]}"
            if is_block:
                text += f"\n<b>Дата блокировки</b> - {datetime.fromisoformat(db.get_block_datetime(user[0])).strftime('%d.%m.%Y %H:%M:%S')}\n\n<i>🙅‍♂️ Пользователь заблокирован</i>"
            else:
                text += f"\n<b>Дата разблокировки</b> - {datetime.fromisoformat(db.get_block_datetime(user[0])).strftime('%d.%m.%Y %H:%M:%S')}\n\n<i>🤵 Пользователь разблокирован</i>"

            await callback.message.edit_text(text=text, reply_markup=paginator.get_page_keyboard(callback.data, option=is_block))

@dp.callback_query_handler(is_admin_filter, lambda call: call.data.startswith("unblockuser"))
async def unblock_user_item(callback: types.CallbackQuery, state: FSMContext):
    async with state.proxy() as data:
        logging.info(f"{sys._getframe().f_code.co_name} | Админ; {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {data.as_dict() = }")
        paginator = data['item_block_users']['paginator']

    user = paginator.get_data()[int(callback.data.split("_")[-1])]
    
    if user[3]:
        user_name, user_username = db.get_names_by_id(user[3])
        username_text = f"@{user_username}" if user_username else ''
        db.del_block_by_id(user[0], datetime.now())

        await callback.message.edit_text(text=f"<b>ID</b> - {user[4]}\n<b>Username</b> - {username_text}\n<b>Имя</b> - {user_name}\n<b>Дата регистрации</b> - {datetime.fromisoformat(user[-1]).strftime('%d.%m.%Y %H:%M:%S')}\n<b>Дата разблокировки</b> - {datetime.fromisoformat(db.get_block_datetime(user[0])).strftime('%d.%m.%Y %H:%M:%S')}\n\n<i>🤵 Пользователь разблокирован</i>", reply_markup=paginator.get_page_keyboard(callback.data, option=False))
    else:
        db.del_block_by_id(user[0], datetime.now())
        user_id = db.get_block_user_id(user[0])
        if user_id:
            _, user_chat_id, user_username, user_name, user_register_datetime = db.get_user_info_by_id(user_id)
            username_text = f"@{user_username}" if user_username else ''

            await callback.message.edit_text(text=f"<b>ID</b> - {user_chat_id}\n<b>Username</b> - {username_text}\n<b>Имя</b> - {user_name}\n<b>Дата регистрации</b> - {datetime.fromisoformat(user_register_datetime).strftime('%d.%m.%Y %H:%M:%S')}\n<b>Дата разблокировки</b> - {datetime.fromisoformat(db.get_block_datetime(user[0])).strftime('%d.%m.%Y %H:%M:%S')}\n\n<i>🤵 Пользователь разблокирован</i>", reply_markup=paginator.get_page_keyboard(callback.data, option=False))
        else:
            await callback.message.edit_text(text=f"<b>Username</b> - {user[1]}\n<b>Дата разблокировки</b> - {datetime.fromisoformat(db.get_block_datetime(user[0])).strftime('%d.%m.%Y %H:%M:%S')}\n\n<i>🤵 Пользователь разблокирован</i>", reply_markup=paginator.get_page_keyboard(callback.data, option=False))

@dp.callback_query_handler(is_admin_filter, lambda call: call.data.startswith("blockuser"))
async def block_user_item(callback: types.CallbackQuery, state: FSMContext):
    async with state.proxy() as data:
        logging.info(f"{sys._getframe().f_code.co_name} | Админ; {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {data.as_dict() = }")
        paginator = data['item_block_users']['paginator']

    user = paginator.get_data()[int(callback.data.split("_")[-1])]
    is_block = bool(db.is_user_block(user[0]))

    if user[3]:
        user_name, user_username = db.get_names_by_id(user[3])
        username_text = f"@{user_username}" if user_username else ''
        if not is_block:
            db.update_block(user[0], datetime.now())

        await callback.message.edit_text(text=f"<b>ID</b> - {user[4]}\n<b>Username</b> - {username_text}\n<b>Имя</b> - {user_name}\n<b>Дата регистрации</b> - {datetime.fromisoformat(user[-1]).strftime('%d.%m.%Y %H:%M:%S')}\n<b>Дата блокировки</b> - {datetime.fromisoformat(db.get_block_datetime(user[0])).strftime('%d.%m.%Y %H:%M:%S')}\n\n<i>🙅‍♂️ Пользователь заблокирован</i>", reply_markup=paginator.get_page_keyboard(callback.data, option=True))
    else:
        user_id = db.get_block_user_id(user[0])
        if user_id:
            _, user_chat_id, user_username, user_name, user_register_datetime = db.get_user_info_by_id(user_id)
            username_text = f"@{user_username}" if user_username else ''
            if not is_block:
                db.update_block(user[0], datetime.now())

            await callback.message.edit_text(text=f"<b>ID</b> - {user_chat_id}\n<b>Username</b> - {username_text}\n<b>Имя</b> - {user_name}\n<b>Дата регистрации</b> - {datetime.fromisoformat(user_register_datetime).strftime('%d.%m.%Y %H:%M:%S')}\n<b>Дата блокировки</b> - {datetime.fromisoformat(db.get_block_datetime(user[0])).strftime('%d.%m.%Y %H:%M:%S')}\n\n<i>🙅‍♂️ Пользователь заблокирован</i>", reply_markup=paginator.get_page_keyboard(callback.data, option=True))
        else:
            if not is_block:
                db.update_block(user[0], datetime.now())
            await callback.message.edit_text(text=f"<b>Username</b> - {user[1]}\n<b>Дата блокировки</b> - {datetime.fromisoformat(db.get_block_datetime(user[0])).strftime('%d.%m.%Y %H:%M:%S')}\n\n<i>🙅‍♂️ Пользователь заблокирован</i>", reply_markup=paginator.get_page_keyboard(callback.data, option=True))

@dp.callback_query_handler(is_admin_filter, text="block_user")
async def block_user(callback: types.CallbackQuery, state: FSMContext):
    logging.info(f"{sys._getframe().f_code.co_name} | Админ; {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {await state.get_data() = }")
    await UnBlockUser.start(callback, state, True)
    
@dp.callback_query_handler(is_admin_filter, text="unblock_user")
async def unblock_user(callback: types.CallbackQuery, state: FSMContext):
    logging.info(f"{sys._getframe().f_code.co_name} | Админ; {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {await state.get_data() = }")
    await UnBlockUser.start(callback, state, False)

class UnBlockUser:
    @staticmethod
    async def cancel(message: types.Message, state: FSMContext):
        logging.debug(f"UnBlockUser.cancel | {message.chat.id = }; {message.from_user.username = }; {await state.get_state() = }; {await state.get_data() = }")
        if message.text == "✖️ Отмена":
            logging.info(f"UnBlockUser.cancel | Пользователь нажал \"Отмена\"; {message.chat.id = }; {message.from_user.username = }; {await state.get_state() = }; {await state.get_data() = }")
            await state.finish()

            message_d = await message.answer(text="🕐 Загрузка ...", reply_markup=remove_markup)
            await message_d.delete()

            await message.answer("Выберите <b>пункт</b>:", reply_markup=users_markup)
            await send_mailing(message.chat.id)

            return False
        return True

    @staticmethod
    async def start(callback: types.CallbackQuery, state: FSMContext, block):
        logging.info(f"UnBlockUser.start | {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }; {block = }; {await state.get_state() = }; {await state.get_data() = }")
        await UnBlockUserStatesGroup.first()
        await state.update_data(block=block)
        
        await callback.message.delete()
        await callback.message.answer(text="Введите <b>username пользователя</b>:", reply_markup=cancel_markup)
        db.update_end_message(0, callback.message.chat.id)

    @staticmethod
    @dp.message_handler(is_admin_filter, content_types=["text"], state=UnBlockUserStatesGroup.get_username)
    async def get_username(message: types.Message, state: FSMContext):
        logging.info(f"{sys._getframe().f_code.co_name} | Админ; {message.chat.id = }; {message.from_user.username = }; {await state.get_state() = }; {await state.get_data() = }")
        if await UnBlockUser.cancel(message, state):
            data = await state.get_data()
            block = data["block"]
            await state.finish()
            message_d = await message.answer(text="🕐 Загрузка ...", reply_markup=remove_markup)
            await message_d.delete()

            username = message.text.strip().strip("@")
            if (username, ) in db.get_block_usernames():
                if block:
                    await message.answer("Этот пользователь <b>УЖЕ</b> <i>заблокирован</i>", reply_markup=users_markup)
                    return
            else:
                if not block:
                    await message.answer("Этот пользователь <b>НЕ</b> <i>заблокирован</i>", reply_markup=users_markup)
                    return

            ids = db.get_user_info_by_username(username)

            if ids:
                user_id, chat_id, user_name, user_register_datetime = ids
                if int(chat_id) in admins:
                    if block:
                        await message.answer("Вы не можете <i>заблокировать</i> <b>админа</b>", reply_markup=users_markup)
                else:
                    if block:
                        logging.info(f"get_username | Блокировка пользователя - {username = }; {ids = }. Инициатор - {message.chat.id = }; {message.from_user.username = }")
                        if db.is_user_in_block(username):
                            db.update_block_by_username(username, datetime.now())
                        else:
                            db.add_block(user_id, username, datetime.now())
                        await message.answer(f"<b>ID</b> - {chat_id}\n<b>Username</b> - @{username}\n<b>Имя</b> - {user_name}\n<b>Дата регистрации</b> - {datetime.fromisoformat(user_register_datetime).strftime('%d.%m.%Y %H:%M:%S')}\n\nПользователь с данными выше <i>заблокирован</i>", reply_markup=users_markup)
                    else:
                        logging.info(f"get_username | Разлокировка пользователя - {username = }; {ids = }. Инициатор - {message.chat.id = }; {message.from_user.username = }")
                        db.del_block_by_username(username, datetime.now())
                        await message.answer(f"<b>ID</b> - {chat_id}\n<b>Username</b> - @{username}\n<b>Имя</b> - {user_name}\n<b>Дата регистрации</b> - {datetime.fromisoformat(user_register_datetime).strftime('%d.%m.%Y %H:%M:%S')}\n\nПользователь с данными выше <i>разблокирован</i>", reply_markup=users_markup)
            else:
                if block:
                    logging.info(f"get_username | Блокировка пользователя - {username = }; {ids = }. Инициатор - {message.chat.id = }; {message.from_user.username = }")
                    if db.is_user_in_block(username):
                        db.update_block_by_username(username, datetime.now())
                    else:
                        db.add_block(username=username, block_datetime=datetime.now())
                    await message.answer(f"<b>Username</b> - @{username}\n\nПользователь с данными выше <i>заблокирован</i>", reply_markup=users_markup)
                else:
                    logging.info(f"get_username | Разлокировка пользователя - {username = }; {ids = }. Инициатор - {message.chat.id = }; {message.from_user.username = }")
                    db.del_block_by_username(username, datetime.now())
                    await message.answer(f"<b>Username</b> - @{username}\n\nПользователь с данными выше <i>разблокирован</i>", reply_markup=users_markup)
            await send_mailing(message.chat.id)

@dp.callback_query_handler(is_admin_filter, text="logs")
async def logs(callback: types.CallbackQuery):
    logging.info(f"{sys._getframe().f_code.co_name} | Админ; {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }")
    
    reply_markup = types.InlineKeyboardMarkup(row_width=2)
    reply_markup.add(*[types.InlineKeyboardButton(file, callback_data=f"logs_{file}") for file in sorted(os.listdir(path_to_log_dir), key=lambda f: 0 if len(f.split(".")) == 2 else int(f.split(".")[-1]))], types.InlineKeyboardButton("все логи", callback_data="logs_all"))
    reply_markup.add(InlineKeyboardButton(text="🔙 Назад", callback_data="admin_panel"))
    await callback.message.edit_text("Выберите <b>файл с логами</b>:", reply_markup=reply_markup)

@dp.callback_query_handler(is_admin_filter, lambda call: call.data.startswith("logs"))
async def log_file(callback: types.CallbackQuery):
    logging.info(f"{sys._getframe().f_code.co_name} | Админ; {callback.message.chat.id = }; {callback.from_user.username = }; {callback.data = }")

    file = callback.data.split("_", maxsplit=1)[-1]

    db.update_end_message(1, callback.from_user.id)
    if file == "all":
        for file_ in sorted(os.listdir(path_to_log_dir), key=lambda f: 0 if len(f.split(".")) == 2 else int(f.split(".")[-1])):
            input_file = types.InputFile(os.path.join(path_to_log_dir, file_))
            await callback.message.answer_document(document=input_file, caption="Логи")
    else:
        input_file = types.InputFile(os.path.join(path_to_log_dir, file))
        await callback.message.answer_document(document=input_file, caption="Логи")

@dp.message_handler(is_not_block_filter, content_types=['text'])
async def text(message: types.Message):
    logging.info(f"{sys._getframe().f_code.co_name} | {message.chat.id = }; {message.from_user.username = }")

    db.update_end_message(0, message.chat.id)
    await message.answer("❌ <i>Команда не распознана</i>.\n\nНажмите <b>/start</b> чтобы вызвать <i>меню</i>")

async def send_mailing(chat_id):
    if chat_id in admins:
        messages = db.get_ends_of_mailings(chat_id)

        logging.info(f"send_mailing | Админ; {chat_id = }; {messages = }")

        for message in messages:
            await bot.send_message(chat_id, message[1])
            db.del_end_of_mailing(message[0])
        
        if messages:
            db.update_end_message(1, chat_id)
    else:
        user_id, datetime = db.get_user_id_and_datetime_with_chat_id(chat_id)
        mailing = db.get_not_send_mailing(user_id, datetime)

        logging.info(f"send_mailing | Не админ; {chat_id = }; {user_id = }; {mailing = }")

        if mailing:
            await bot.copy_message(chat_id, mailing[2], mailing[1])
            db.add_mailings_users_with_mailing_id_and_user_id(mailing[0], user_id)
            db.update_success_count_in_mailing(mailing[1])
            db.update_end_message(1, chat_id)


if __name__ == '__main__':
    if not os.path.exists(path_to_data_dir): os.mkdir(path_to_data_dir)
    if not os.path.exists(path_to_log_dir): os.mkdir(path_to_log_dir)

    file_log = RotatingFileHandler(os.path.join(path_to_log_dir, "calories_bot.log"), maxBytes=5_000_000, backupCount=10, encoding="utf8")
    console_out = logging.StreamHandler()

    logging.basicConfig(handlers=(file_log, console_out), 
                        format='[%(asctime)s | %(levelname)s]: %(message)s', 
                        datefmt='%m.%d.%Y %H:%M:%S',
                        level=logging.INFO)

    scheduler.add_job(lambda: db.clear_user_activities(datetime.now() - timedelta(days=60)), trigger='cron', hour=1)
    scheduler.start()

    PaginatorDiary.init("diary", "product_calories")

    executor.start_polling(dp, skip_updates=True, on_startup=on_startup)
